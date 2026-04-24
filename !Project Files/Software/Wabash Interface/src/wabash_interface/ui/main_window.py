from __future__ import annotations

import os
import sys
import json
import csv
import time
import datetime
import subprocess
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
import serial
from tksheet import Sheet

from serial.tools import list_ports

from wabash_interface.services.serial_service import SerialConfig, SerialService
from wabash_interface.storage.log_export import export_text_log

WABASH_BLUE        = "#1676D2"
WABASH_BLUE_HOVER  = "#105DA8"
BTN_GREY          = "#334E73"
BTN_GREY_HOVER    = "#293F5C"
PANEL_LIGHT       = "#F2F6FB"
PANEL_DARK        = "#0D1726"
CARD_LIGHT        = "#FFFFFF"
CARD_DARK         = "#162338"

# Log line colour tags (foreground only; works on both light/dark)
LOG_COLORS = {
    "tx":      "#60A5FA",   # bright blue â€” sent commands
    "rx":      "#93C5FD",   # soft blue   â€” received data
    "data":    "#34D399",   # green     â€” DATA/DATC payload rows
    "status":  "#FBBF24",   # amber     â€” status/RSP lines
    "default": "#D1D5DB",   # grey      â€” everything else
}

# Density presets: (pad_x, pad_y_btn, font_size_log, row_padding)
DENSITY = {
    "Compact":     (8,  3, 11, 4),
    "Comfortable": (16, 6, 12, 8),
}

SETUP_MASK_SENSOR_INTERVAL = 1 << 0
SETUP_MASK_THRESHOLD = 1 << 1
SETUP_MASK_SAMPLE_RATE = 1 << 2
SETUP_MASK_DURATION = 1 << 3
SETUP_MASK_TRUCK_ID = 1 << 4
SETUP_MASK_DESCRIPTION = 1 << 5
SETUP_MASK_WIFI = 1 << 6

OFFLOAD_STATUS_COLORS = {
    "Idle": ("#64748B", "#94A3B8"),
    "Connecting...": (WABASH_BLUE, "#60A5FA"),
    "Transferring...": ("#0F766E", "#2DD4BF"),
    "Complete": ("#15803D", "#4ADE80"),
    "Timeout": ("#B45309", "#FBBF24"),
    "Failed": ("#DC2626", "#F87171"),
    "Fallback LoRa": ("#7C3AED", "#C4B5FD"),
    "No Data": ("#475569", "#CBD5E1"),
}


def _asset(rel: str) -> Path:
    """Return path to a bundled asset whether running frozen or from source."""
    if getattr(sys, 'frozen', False):
        base = Path(sys._MEIPASS)  # type: ignore[attr-defined]
    else:
        base = Path(__file__).parent.parent.parent.parent  # project root
    return base / rel


# App-local storage (Windows: %LOCALAPPDATA%\WabashInterface)
_APP_DIR: Path = Path(
    os.environ.get("LOCALAPPDATA") or (Path.home() / "AppData" / "Local")
) / "WabashInterface"
_SETTINGS_FILE: Path = _APP_DIR / "settings.json"
_KNOWN_TRUCKS_FILE: Path = _APP_DIR / "known_trucks.txt"


def _default_data_path() -> Path:
    return _APP_DIR / "data"


def _excel_column_name(index: int) -> str:
    """Convert a zero-based column index to an Excel-style column label."""
    label = ""
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        label = chr(65 + remainder) + label
    return label


class MainWindow:
    def __init__(self) -> None:
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        self.root = ctk.CTk()
        self.root.title("Wabash Interface")
        self.root.geometry("1240x780")
        self.root.minsize(1024, 680)
        self.root.after(0, lambda: self.root.state("zoomed"))

        self.serial_service = SerialService()
        self.log_lines: list[str] = []
        self.tx_count = 0
        self.rx_count = 0
        self.connected = False
        self.connected_role = "disconnected"
        self.last_message = "No messages yet"
        self.lora_active = False
        self.session_events = 0
        self.units: dict[str, dict] = {}
        self.known_truck_ids: set[str] = set()
        self.detected_port_roles: dict[str, str] = {}
        self.connected_port = "-"
        self._auto_connect_inflight = False
        self._search_highlight_job: str | None = None
        self._wrap_job: str | None = None
        self._resize_job: str | None = None
        self._resizing: bool = False
        self._last_size: tuple[int, int] = (0, 0)
        self._next_offload_ui_update: float = 0.0
        self._window_interacting_until: float = 0.0

        # Offload statistics tracking
        self.offload_in_progress = False
        self.offload_start_time: float | None = None
        self.offload_wifi_start_time: float | None = None
        self.offload_connected_time: float | None = None
        self.offload_data_start_time: float | None = None
        self.offload_end_time: float | None = None
        self.offload_connection_duration: float | None = None
        self.offload_transfer_duration: float | None = None
        self.offload_events_count = 0
        self.offload_last_status = "Idle"
        self.offload_status = "Idle"  # "Idle", "Connecting...", "Transferring...", "Complete", "Failed", "Timeout"

        # Per-event file saving state
        self._offload_pending_event_name: str | None = None
        self._offload_pending_rows: list[str] = []
        self._offload_lora_row_buf: str = ""
        self._offload_saved_files: list[Path] = []
        self._offload_session_dir: Path | None = None
        self._offload_summary_file: Path | None = None

        # Data storage path (defaults to AppData, overridden by saved settings)
        self.data_path_var = tk.StringVar(value=str(_default_data_path()))
        self._load_app_settings()
        self._load_known_units()

        self.port_var     = tk.StringVar(value="")
        self.baud_var     = tk.StringVar(value="115200")
        self.command_var  = tk.StringVar(value="")
        self.viewer_sheet_var = tk.StringVar(value="")
        self.viewer_filter_var = tk.StringVar(value="")
        self.theme_var    = tk.StringVar(value="system")
        self.density_var  = tk.StringVar(value="Comfortable")
        self.search_var      = tk.StringVar(value="")
        self.unit_filter_var = tk.StringVar(value="")
        self.scan_var        = tk.StringVar(value="No units found")
        self.text_scale_var  = tk.DoubleVar(value=1.0)
        self.search_placeholder_active = False
        self._scan_results: list[str] = []

        # Unit Setup configuration variables
        self.sensor_interval_var = tk.StringVar(value="100")
        self.event_trigger_threshold_var = tk.StringVar(value="2.0")
        self.lab_sample_rate_var = tk.StringVar(value="20")
        self.event_duration_var = tk.StringVar(value="2000")
        self.truck_id_var = tk.StringVar(value="")
        self.description_var = tk.StringVar(value="")

        self.apply_sensor_interval_var = tk.BooleanVar(value=False)
        self.apply_threshold_var = tk.BooleanVar(value=False)
        self.apply_sample_rate_var = tk.BooleanVar(value=False)
        self.apply_duration_var = tk.BooleanVar(value=False)
        self.apply_truck_id_var = tk.BooleanVar(value=False)
        self.apply_description_var = tk.BooleanVar(value=False)
        self.apply_wifi_var = tk.BooleanVar(value=False)

        # Wi-Fi credential slot (sent to receiver for Wi-Fi-first offload)
        self.wifi1_ssid_var = tk.StringVar(value="")
        self.wifi1_password_var = tk.StringVar(value="")

        self.send_config_button: ctk.CTkButton | None = None
        self.unit_setup_help_label: ctk.CTkLabel | None = None
        self.unit_setup_desc_card: ctk.CTkFrame | None = None
        self._setup_apply_vars = [
            self.apply_sensor_interval_var,
            self.apply_threshold_var,
            self.apply_sample_rate_var,
            self.apply_duration_var,
            self.apply_truck_id_var,
            self.apply_description_var,
            self.apply_wifi_var,
        ]
        for var in self._setup_apply_vars:
            var.trace_add("write", self._on_setup_selection_changed)

        self.pages: dict[str, ctk.CTkFrame | ctk.CTkScrollableFrame] = {}
        self.nav_buttons: dict[str, ctk.CTkButton] = {}
        self._active_page_name: str | None = None
        self._viewer_workbook_path: Path | None = None
        self._viewer_sheet_rows: list[list[str]] = []
        self._viewer_filtered_rows: list[list[str]] = []
        self._viewer_header_row: list[str] = []
        self._viewer_col_start = 0
        self._viewer_col_span = 10
        self._viewer_sort_column: int | None = None
        self._viewer_sort_descending = False
        self.viewer_tree: ttk.Treeview | None = None
        self.viewer_sheet: Sheet | None = None
        self.viewer_grid_card: ctk.CTkFrame | None = None
        self.viewer_stats_card: ctk.CTkFrame | None = None
        self.viewer_kpi_labels: dict[str, ctk.CTkLabel] = {}
        self.viewer_tree_xscroll: tk.Scrollbar | None = None
        self.viewer_status_label: ctk.CTkLabel | None = None
        self.viewer_source_label: ctk.CTkLabel | None = None
        self.viewer_path_label: ctk.CTkLabel | None = None
        self.viewer_meta_label: ctk.CTkLabel | None = None
        self.viewer_sheet_info_label: ctk.CTkLabel | None = None
        self.viewer_col_window_label: ctk.CTkLabel | None = None
        self.viewer_hint_label: ctk.CTkLabel | None = None
        self.viewer_sheet_combo: ctk.CTkComboBox | None = None
        self._viewer_detail_window: ctk.CTkToplevel | None = None
        self._viewer_strain_low_thresh: float = 0.0
        self._viewer_strain_high_thresh: float = 0.0
        self._viewer_max_strain_row_idx: int = -1
        self._viewer_max_accel_row_idx: int = -1
        self._viewer_max_strain_col_idx: int = -1
        self._viewer_max_accel_col_idx: int = -1
        self.viewer_filter_var.trace_add("write", lambda *_: self._apply_viewer_filter())

        # set window icon
        icon_path = _asset("assets/images/icon.ico")
        if icon_path.exists():
            self.root.iconbitmap(str(icon_path))

        self._build_ui()
        self._refresh_unit_list()
        self._show_page("Dashboard")
        self._refresh_ports()
        self._pump_messages()
        self.root.after(350, self._auto_connect_tick)

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        # Suppress per-pixel geometry recalculations while the window is being dragged.
        self.root.bind("<Configure>", self._on_root_configure, add="+")

    def run(self) -> None:
        self.root.mainloop()

    def _build_ui(self) -> None:
        self.root.configure(fg_color=(PANEL_LIGHT, PANEL_DARK))
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        sidebar = ctk.CTkFrame(self.root, width=280, corner_radius=0, fg_color=("#E5E7EB", "#0B1220"))
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        sidebar.grid_columnconfigure(0, weight=1)

        self.page_container = ctk.CTkFrame(self.root, corner_radius=0, fg_color="transparent")
        self.page_container.grid(row=0, column=1, sticky="nsew", padx=14, pady=14)
        self.page_container.grid_propagate(False)
        self.page_container.grid_columnconfigure(0, weight=1)
        self.page_container.grid_rowconfigure(0, weight=1)

        ctk.CTkLabel(
            sidebar,
            text="WABASH",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=WABASH_BLUE,
        ).grid(row=0, column=0, sticky="w", padx=24, pady=(28, 4))
        ctk.CTkLabel(
            sidebar,
            text="Receiver Command Center",
            text_color=("#475569", "#9CA3AF"),
            font=ctk.CTkFont(size=14),
        ).grid(row=1, column=0, sticky="w", padx=24, pady=(0, 20))

        ctk.CTkFrame(sidebar, height=3, corner_radius=8, fg_color=WABASH_BLUE).grid(
            row=1, column=0, sticky="ew", padx=24, pady=(18, 0)
        )

        nav_card = ctk.CTkFrame(sidebar, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        nav_card.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 12))
        nav_card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(nav_card, text="Navigation", font=ctk.CTkFont(size=16, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 8)
        )

        self.nav_buttons["Dashboard"] = ctk.CTkButton(
            nav_card,
            text="Dashboard",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=lambda: self._show_page("Dashboard"),
        )
        self.nav_buttons["Dashboard"].grid(row=1, column=0, sticky="ew", padx=16, pady=4)

        self.nav_buttons["Unit Setup"] = ctk.CTkButton(
            nav_card,
            text="Unit Setup",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=lambda: self._show_page("Unit Setup"),
        )
        self.nav_buttons["Unit Setup"].grid(row=2, column=0, sticky="ew", padx=16, pady=4)

        self.nav_buttons["Data Offload"] = ctk.CTkButton(
            nav_card,
            text="Data Offload",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=lambda: self._show_page("Data Offload"),
        )
        self.nav_buttons["Data Offload"].grid(row=3, column=0, sticky="ew", padx=16, pady=4)

        self.nav_buttons["Data Viewer"] = ctk.CTkButton(
            nav_card,
            text="Data Viewer",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=lambda: self._show_page("Data Viewer"),
        )
        self.nav_buttons["Data Viewer"].grid(row=4, column=0, sticky="ew", padx=16, pady=4)

        self.nav_buttons["Settings"] = ctk.CTkButton(
            nav_card,
            text="Settings",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=lambda: self._show_page("Settings"),
        )
        self.nav_buttons["Settings"].grid(row=5, column=0, sticky="ew", padx=16, pady=(4, 14))

        quick_status = ctk.CTkFrame(sidebar, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        quick_status.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 18))
        quick_status.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(quick_status, text="System Status", font=ctk.CTkFont(size=16, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 8)
        )
        self.sidebar_status = ctk.CTkLabel(
            quick_status,
            text="Disconnected",
            corner_radius=10,
            fg_color="#7F1D1D",
            text_color="#FEE2E2",
        )
        self.sidebar_status.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 14))

        self.sidebar_board_label = ctk.CTkLabel(
            quick_status,
            text="Selected Board: None",
            text_color=("#475569", "#94A3B8"),
            justify="left",
            anchor="w",
            wraplength=220,
            font=ctk.CTkFont(size=12),
        )
        self.sidebar_board_label.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 10))

        quick_btn_row = ctk.CTkFrame(quick_status, fg_color="transparent")
        quick_btn_row.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 8))
        quick_btn_row.grid_columnconfigure((0, 1), weight=1)
        self.sidebar_connect_button = ctk.CTkButton(
            quick_btn_row,
            text="Connect",
            fg_color=WABASH_BLUE,
            hover_color=WABASH_BLUE_HOVER,
            command=self._connect,
        )
        self.sidebar_connect_button.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        self.sidebar_disconnect_button = ctk.CTkButton(
            quick_btn_row,
            text="Disconnect",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=self._disconnect,
        )
        self.sidebar_disconnect_button.grid(row=0, column=1, sticky="ew", padx=(4, 0))

        self.sidebar_auto_detect_button = ctk.CTkButton(
            quick_status,
            text="Auto Detect Transmitter",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=self._auto_detect_board,
        )
        self.sidebar_auto_detect_button.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 14))

        self._build_dashboard_page()
        self._build_settings_page()
        self._build_live_page()
        self._build_data_viewer_page()
        self._build_unit_setup_page()

    def _create_stat_card(self, parent: ctk.CTkFrame, column: int, title: str, initial: str) -> ctk.CTkLabel:
        card = ctk.CTkFrame(parent, corner_radius=12, fg_color=(CARD_LIGHT, CARD_DARK))
        card.grid(row=0, column=column, sticky="ew", padx=(0 if column == 0 else 6, 0 if column == 2 else 6))
        ctk.CTkLabel(card, text=title, text_color=("#475569", "#94A3B8"), font=ctk.CTkFont(size=12)).grid(
            row=0, column=0, sticky="w", padx=14, pady=(10, 0)
        )
        value = ctk.CTkLabel(card, text=initial, font=ctk.CTkFont(size=22, weight="bold"))
        value.grid(row=1, column=0, sticky="w", padx=14, pady=(2, 10))
        return value

    def _show_page(self, page_name: str) -> None:
        if self._active_page_name == page_name:
            return

        if self._active_page_name is not None and self._active_page_name in self.pages:
            self.pages[self._active_page_name].grid_remove()

        self.pages[page_name].grid(row=0, column=0, sticky="nsew")
        self._active_page_name = page_name

        if page_name == "Data Viewer":
            self._refresh_data_viewer()

        for name, button in self.nav_buttons.items():
            if name == page_name:
                button.configure(fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER)
            else:
                button.configure(fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER)

    def _build_dashboard_page(self) -> None:
        page = ctk.CTkFrame(self.page_container, corner_radius=0, fg_color="transparent")
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(2, weight=1)
        self.pages["Dashboard"] = page

        # ── Header ──────────────────────────────────────────────────────────
        header = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        header.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        header.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(header, text="Fleet Dashboard", font=ctk.CTkFont(size=22, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=18, pady=14
        )
        ctk.CTkLabel(
            header, text="WABASH", text_color=WABASH_BLUE, font=ctk.CTkFont(size=14, weight="bold")
        ).grid(row=0, column=1, sticky="e", padx=18, pady=14)

        # ── System Status Cards ──────────────────────────────────────────────
        status_row = ctk.CTkFrame(page, fg_color="transparent")
        status_row.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        status_row.grid_columnconfigure((0, 1, 2), weight=1)

        tx_card = ctk.CTkFrame(status_row, corner_radius=12, fg_color=(CARD_LIGHT, CARD_DARK))
        tx_card.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        tx_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            tx_card, text="Transmitter", text_color=("#64748B", "#94A3B8"), font=ctk.CTkFont(size=12)
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(10, 0))
        self.db_tx_status = ctk.CTkLabel(
            tx_card, text="Not Connected",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=("#EF4444", "#EF4444"),
        )
        self.db_tx_status.grid(row=1, column=0, sticky="w", padx=14, pady=(2, 10))

        lora_card = ctk.CTkFrame(status_row, corner_radius=12, fg_color=(CARD_LIGHT, CARD_DARK))
        lora_card.grid(row=0, column=1, sticky="ew", padx=6)
        lora_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            lora_card, text="Truck / LoRa Link", text_color=("#64748B", "#94A3B8"), font=ctk.CTkFont(size=12)
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(10, 0))
        self.db_lora_status = ctk.CTkLabel(
            lora_card, text="No Link",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=("#F59E0B", "#F59E0B"),
        )
        self.db_lora_status.grid(row=1, column=0, sticky="w", padx=14, pady=(2, 10))

        sess_card = ctk.CTkFrame(status_row, corner_radius=12, fg_color=(CARD_LIGHT, CARD_DARK))
        sess_card.grid(row=0, column=2, sticky="ew", padx=(6, 0))
        sess_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            sess_card, text="Events This Session", text_color=("#64748B", "#94A3B8"), font=ctk.CTkFont(size=12)
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(10, 0))
        self.db_event_count = ctk.CTkLabel(sess_card, text="0", font=ctk.CTkFont(size=22, weight="bold"))
        self.db_event_count.grid(row=1, column=0, sticky="w", padx=14, pady=(2, 10))

        # ── Main row: fleet table + active unit panel ────────────────────────
        main_row = ctk.CTkFrame(page, fg_color="transparent")
        main_row.grid(row=2, column=0, sticky="nsew", pady=(0, 12))
        main_row.grid_columnconfigure(0, weight=1)
        main_row.grid_rowconfigure(0, weight=1)

        # Fleet table (left)
        fleet_card = ctk.CTkFrame(main_row, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        fleet_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        fleet_card.grid_columnconfigure(0, weight=1)
        fleet_card.grid_rowconfigure(2, weight=1)

        fleet_hdr = ctk.CTkFrame(fleet_card, fg_color="transparent")
        fleet_hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 6))
        fleet_hdr.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(fleet_hdr, text="Known Units", font=ctk.CTkFont(size=16, weight="bold")).grid(
            row=0, column=0, sticky="w"
        )
        self.unit_filter_entry = ctk.CTkEntry(
            fleet_hdr, textvariable=self.unit_filter_var, placeholder_text="Filter by ID...", width=180
        )
        self.unit_filter_entry.grid(row=0, column=1, sticky="e")
        self.unit_filter_var.trace_add("write", lambda *_: self._refresh_unit_list())

        fleet_col_mins = (120, 98, 64, 76, 74)
        fleet_col_weights = (4, 3, 2, 2, 2)

        # Single-column scroll frame: row frames fill the full width  →  smooth backgrounds.
        # Header frame and row frames share identical internal 5-column layout  →  perfect alignment.
        self.fleet_scroll = ctk.CTkScrollableFrame(fleet_card, corner_radius=8, fg_color="transparent")
        self.fleet_scroll.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 14))
        self.fleet_scroll.grid_columnconfigure(0, weight=1)

        self.fleet_thead = ctk.CTkFrame(self.fleet_scroll, fg_color=("#DCE5F5", "#22365A"), corner_radius=8)
        self.fleet_thead.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        for col, (min_w, weight) in enumerate(zip(fleet_col_mins, fleet_col_weights)):
            self.fleet_thead.grid_columnconfigure(col, minsize=min_w, weight=weight)
        for col, lbl_text in enumerate(["Truck ID", "Last Seen", "Events", "Status", "Action"]):
            _a = "e" if col == 4 else "w"
            ctk.CTkLabel(
                self.fleet_thead, text=lbl_text,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color=("#334155", "#BFDBFE"),
                anchor=_a,
            ).grid(row=0, column=col, sticky="ew", padx=12, pady=8)

        self.fleet_empty_label = ctk.CTkLabel(
            self.fleet_scroll,
            text="No units known yet.\nConnect and offload data to populate this list.",
            text_color=("#64748B", "#6B7280"),
            font=ctk.CTkFont(size=13),
            justify="center",
        )
        self.fleet_empty_label.grid(row=1, column=0, pady=30)
        self._fleet_rows: list = []

        # Active unit panel (right)
        unit_card = ctk.CTkFrame(main_row, width=260, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        unit_card.grid(row=0, column=1, sticky="nsew")
        unit_card.grid_propagate(False)
        unit_card.grid_columnconfigure(0, weight=1)
        unit_card.grid_rowconfigure(12, weight=1)

        ctk.CTkLabel(unit_card, text="Active Unit", font=ctk.CTkFont(size=16, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 8)
        )
        ctk.CTkFrame(unit_card, height=1, fg_color=("#E5E7EB", "#1E3050")).grid(
            row=1, column=0, sticky="ew", padx=16
        )
        self.db_unit_id = ctk.CTkLabel(
            unit_card, text="No unit configured",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=("#64748B", "#9CA3AF"),
            wraplength=220, justify="left",
        )
        self.db_unit_id.grid(row=2, column=0, sticky="w", padx=16, pady=(10, 2))
        self.db_unit_desc = ctk.CTkLabel(
            unit_card, text="",
            font=ctk.CTkFont(size=12),
            text_color=("#64748B", "#9CA3AF"),
            wraplength=220, justify="left",
        )
        self.db_unit_desc.grid(row=3, column=0, sticky="w", padx=16, pady=(0, 4))

        ctk.CTkFrame(unit_card, height=1, fg_color=("#E5E7EB", "#1E3050")).grid(
            row=4, column=0, sticky="ew", padx=16, pady=(8, 0)
        )
        ctk.CTkLabel(
            unit_card, text="Session Stats",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=("#475569", "#94A3B8"),
        ).grid(row=5, column=0, sticky="w", padx=16, pady=(10, 4))

        for stat_i, (stat_label, attr_name) in enumerate([
            ("Lines Received", "db_rx_count"),
            ("Commands Sent",  "db_tx_count"),
            ("Log Lines",      "db_log_count"),
        ]):
            ctk.CTkLabel(
                unit_card, text=stat_label,
                font=ctk.CTkFont(size=11),
                text_color=("#64748B", "#9CA3AF"),
            ).grid(row=6 + stat_i * 2, column=0, sticky="w", padx=16, pady=(4, 0))
            val_lbl = ctk.CTkLabel(unit_card, text="0", font=ctk.CTkFont(size=16, weight="bold"))
            val_lbl.grid(row=7 + stat_i * 2, column=0, sticky="w", padx=16, pady=(0, 2))
            setattr(self, attr_name, val_lbl)

        ctk.CTkButton(
            unit_card, text="Open Data Offload",
            command=lambda: self._show_page("Data Offload"),
            fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER,
            font=ctk.CTkFont(size=13, weight="bold"),
        ).grid(row=13, column=0, sticky="ew", padx=16, pady=(0, 6))
        ctk.CTkButton(
            unit_card, text="Configure Unit",
            command=lambda: self._show_page("Unit Setup"),
            fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER,
            font=ctk.CTkFont(size=13),
        ).grid(row=14, column=0, sticky="ew", padx=16, pady=(0, 14))

        # ── Latest message strip ─────────────────────────────────────────────
        activity_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        activity_card.grid(row=3, column=0, sticky="ew")
        activity_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(activity_card, text="Latest Message", font=ctk.CTkFont(size=14, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(12, 4)
        )
        self.db_last_message = ctk.CTkLabel(
            activity_card,
            text="No messages yet",
            wraplength=760,
            justify="left",
            anchor="w",
            text_color=("#334155", "#CBD5E1"),
            font=ctk.CTkFont(size=12),
        )
        self.db_last_message.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 12))

    def _build_settings_page(self) -> None:
        page = ctk.CTkFrame(self.page_container, corner_radius=0, fg_color="transparent")
        page.grid_columnconfigure(0, weight=1)
        self.pages["Settings"] = page

        header = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        header.grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(header, text="Settings", font=ctk.CTkFont(size=22, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=18, pady=14
        )

        card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        card.grid(row=1, column=0, sticky="ew", pady=(12, 0))
        card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(card, text="Appearance", font=ctk.CTkFont(size=16, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 6)
        )
        ctk.CTkLabel(card, text="Theme", text_color=("#475569", "#94A3B8")).grid(
            row=1, column=0, sticky="w", padx=16
        )
        ctk.CTkSegmentedButton(
            card,
            values=["dark", "light", "system"],
            variable=self.theme_var,
            command=self._set_theme,
            selected_color=WABASH_BLUE,
            selected_hover_color=WABASH_BLUE_HOVER,
        ).grid(row=2, column=0, sticky="ew", padx=16, pady=(4, 10))

        ctk.CTkLabel(card, text="Layout Density", text_color=("#475569", "#94A3B8")).grid(
            row=3, column=0, sticky="w", padx=16
        )
        ctk.CTkSegmentedButton(
            card,
            values=["Compact", "Comfortable"],
            variable=self.density_var,
            command=self._set_density,
            selected_color=WABASH_BLUE,
            selected_hover_color=WABASH_BLUE_HOVER,
        ).grid(row=4, column=0, sticky="ew", padx=16, pady=(4, 10))

        ctk.CTkLabel(card, text="Text Size", text_color=("#475569", "#94A3B8")).grid(
            row=5, column=0, sticky="w", padx=16
        )
        self.text_scale_label = ctk.CTkLabel(card, text="100%", text_color=("#334155", "#CBD5E1"))
        self.text_scale_label.grid(row=6, column=0, sticky="e", padx=16)
        ctk.CTkSlider(
            card,
            from_=0.9,
            to=1.4,
            number_of_steps=10,
            variable=self.text_scale_var,
            button_color=WABASH_BLUE,
            button_hover_color=WABASH_BLUE_HOVER,
            progress_color=WABASH_BLUE,
            command=self._set_text_scale,
        ).grid(row=7, column=0, sticky="ew", padx=16, pady=(4, 14))

        storage_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        storage_card.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        storage_card.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(
            storage_card, text="Data Storage",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=0, column=0, columnspan=4, sticky="w", padx=16, pady=(14, 6))
        ctk.CTkLabel(
            storage_card, text="Save Location",
            text_color=("#475569", "#94A3B8"),
        ).grid(row=1, column=0, sticky="w", padx=16, pady=6)
        ctk.CTkEntry(
            storage_card,
            textvariable=self.data_path_var,
            state="readonly",
        ).grid(row=1, column=1, sticky="ew", padx=(6, 6), pady=6)
        ctk.CTkButton(
            storage_card,
            text="Open",
            width=70,
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=self._open_data_folder,
        ).grid(row=1, column=2, sticky="e", padx=(0, 6), pady=6)
        ctk.CTkButton(
            storage_card,
            text="Change...",
            width=90,
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=self._browse_data_path,
        ).grid(row=1, column=3, sticky="e", padx=(0, 16), pady=6)
        ctk.CTkLabel(
            storage_card,
            text="Offloaded event files are saved here, organised by Truck ID.",
            text_color=("#475569", "#94A3B8"),
            font=ctk.CTkFont(size=12),
            justify="left",
            anchor="w",
        ).grid(row=2, column=0, columnspan=4, sticky="ew", padx=16, pady=(0, 14))

    def _build_live_page(self) -> None:
        page = ctk.CTkFrame(self.page_container, corner_radius=0, fg_color='transparent')
        page.grid_columnconfigure(0, minsize=340)
        page.grid_columnconfigure(1, weight=1)
        page.grid_rowconfigure(0, weight=1)
        self.pages['Data Offload'] = page

        # ================================================================
        # LEFT PANEL
        # ================================================================
        left = ctk.CTkFrame(page, corner_radius=0,
                            fg_color='transparent')
        left.grid(row=0, column=0, sticky='nsew', padx=(0, 12))
        left.grid_columnconfigure(0, weight=1)
        left.grid_rowconfigure(1, weight=1)

        # Section 1: Transmitter connection
        conn_card = ctk.CTkFrame(left, corner_radius=14,
                                 fg_color=(CARD_LIGHT, CARD_DARK))
        conn_card.grid(row=0, column=0, sticky='ew', padx=12, pady=(0, 8))
        conn_card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(conn_card, text='Transmitter',
                     font=ctk.CTkFont(size=14, weight='bold')).grid(
            row=0, column=0, columnspan=2, sticky='w', padx=14, pady=(12, 6))

        ctk.CTkLabel(conn_card, text='COM Port',
                     text_color=('#475569', '#94A3B8')).grid(
            row=1, column=0, columnspan=2, sticky='w', padx=14)
        self.port_combo = ctk.CTkComboBox(conn_card, variable=self.port_var,
                                          values=['No ports found'])
        self.port_combo.grid(row=2, column=0, columnspan=2, sticky='ew',
                             padx=14, pady=(4, 6))
        self.port_var.trace_add("write", lambda *_: self._update_port_selection_status())

        self.port_detect_label = ctk.CTkLabel(
            conn_card,
            text='Board Detection: Unknown',
            text_color=('#475569', '#94A3B8'),
            justify='left',
            anchor='w',
        )
        self.port_detect_label.grid(row=3, column=0, columnspan=2, sticky='ew', padx=14, pady=(0, 4))

        ctk.CTkLabel(conn_card, text='Baud',
                     text_color=('#475569', '#94A3B8')).grid(
            row=4, column=0, columnspan=2, sticky='w', padx=14)
        ctk.CTkEntry(conn_card, textvariable=self.baud_var).grid(
            row=5, column=0, columnspan=2, sticky='ew', padx=14, pady=(4, 8))

        btn_row = ctk.CTkFrame(conn_card, fg_color='transparent')
        btn_row.grid(row=6, column=0, columnspan=2, sticky='ew',
                     padx=14, pady=(0, 12))
        btn_row.grid_columnconfigure((0, 1, 2), weight=1)
        ctk.CTkButton(btn_row, text='Refresh',
                      fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER,
                      command=self._refresh_ports).grid(
            row=0, column=0, padx=(0, 3), sticky='ew')
        self.connect_button = ctk.CTkButton(
            btn_row, text='Connect',
            fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER,
            command=self._connect)
        self.connect_button.grid(row=0, column=1, padx=3, sticky='ew')
        self.disconnect_button = ctk.CTkButton(
            btn_row, text='Disconnect',
            fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER,
            command=self._disconnect)
        self.disconnect_button.grid(row=0, column=2, padx=(3, 0), sticky='ew')

        # Section 2: Unit Discovery + table (fills remaining height)
        disc_card = ctk.CTkFrame(left, corner_radius=14,
                                 fg_color=(CARD_LIGHT, CARD_DARK))
        disc_card.grid(row=1, column=0, sticky='nsew', padx=12, pady=(0, 8))
        disc_card.grid_columnconfigure(0, weight=1)
        disc_card.grid_rowconfigure(2, weight=1)

        disc_hdr = ctk.CTkFrame(disc_card, fg_color='transparent')
        disc_hdr.grid(row=0, column=0, sticky='ew', padx=14, pady=(12, 8))
        disc_hdr.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(disc_hdr, text='Unit Discovery',
                     font=ctk.CTkFont(size=14, weight='bold')).grid(
            row=0, column=0, sticky='w')
        self.scan_button = ctk.CTkButton(
            disc_hdr, text='Unit Discover', width=110,
            fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER,
            command=self._run_connection_scan)
        self.scan_button.grid(row=0, column=1, sticky='e')

        # Table column header
        thead = ctk.CTkFrame(disc_card, fg_color=('#D1D5DB', '#1E3050'),
                             corner_radius=6)
        thead.grid(row=1, column=0, sticky='ew', padx=14, pady=(0, 2))
        thead.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(thead, text='Truck ID',
                     font=ctk.CTkFont(size=11, weight='bold'),
                     text_color=('#475569', '#94A3B8')).grid(
            row=0, column=0, sticky='w', padx=10, pady=5)
        ctk.CTkLabel(thead, text='Select',
                     font=ctk.CTkFont(size=11, weight='bold'),
                     text_color=('#475569', '#94A3B8')).grid(
            row=0, column=1, sticky='e', padx=10, pady=5)

        # Scrollable table body
        self.disc_scroll = ctk.CTkScrollableFrame(
            disc_card, corner_radius=6, fg_color='transparent')
        self.disc_scroll.grid(row=2, column=0, sticky='nsew', padx=14,
                              pady=(0, 12))
        self.disc_scroll.grid_columnconfigure(0, weight=1)
        self.disc_empty_label = ctk.CTkLabel(
            self.disc_scroll,
            text='No units found.\nPress Unit Discover to scan.',
            text_color=('#64748B', '#6B7280'),
            font=ctk.CTkFont(size=12), justify='center')
        self.disc_empty_label.grid(row=0, column=0, columnspan=2, pady=20)
        self._disc_rows: list = []

        # Section 3: Unit Actions (locked until a unit is selected)
        act_card = ctk.CTkFrame(left, corner_radius=14,
                                fg_color=(CARD_LIGHT, CARD_DARK))
        act_card.grid(row=2, column=0, sticky='ew', padx=12, pady=(0, 0))
        act_card.grid_columnconfigure((0, 1, 2), weight=1)

        ctk.CTkLabel(act_card, text='Unit Actions',
                     font=ctk.CTkFont(size=14, weight='bold')).grid(
            row=0, column=0, columnspan=3, sticky='w', padx=14, pady=(12, 8))

        self.btn_request = ctk.CTkButton(
            act_card, text='Request Data',
            fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER,
            state='disabled',
            command=lambda: self._send_quick('d'))
        self.btn_request.grid(row=1, column=0, padx=(14, 4), pady=(0, 12),
                              sticky='ew')
        self.btn_tare = ctk.CTkButton(
            act_card, text='Unit Tare',
            fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER,
            state='disabled',
            command=lambda: self._send_quick('z'))
        self.btn_tare.grid(row=1, column=1, padx=4, pady=(0, 12), sticky='ew')
        self.btn_timesync = ctk.CTkButton(
            act_card, text='Time Sync',
            fg_color=BTN_GREY, hover_color=BTN_GREY_HOVER,
            state='disabled',
            command=self._send_time_sync)
        self.btn_timesync.grid(row=1, column=2, padx=(4, 14), pady=(0, 12),
                               sticky='ew')

        # ================================================================
        # RIGHT PANEL
        # ================================================================
        right = ctk.CTkFrame(page, corner_radius=0, fg_color='transparent')
        right.grid(row=0, column=1, sticky='nsew')
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(3, weight=1)

        # Header bar (title + connection badge)
        hdr = ctk.CTkFrame(right, corner_radius=14,
                           fg_color=(CARD_LIGHT, CARD_DARK))
        hdr.grid(row=0, column=0, sticky='ew', pady=(0, 10))
        hdr.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(hdr, text='Data Offload',
                     font=ctk.CTkFont(size=22, weight='bold')).grid(
            row=0, column=0, sticky='w', padx=18, pady=14)
        ctk.CTkLabel(hdr, text='WABASH',
                     text_color=WABASH_BLUE,
                     font=ctk.CTkFont(size=16, weight='bold')).grid(
            row=0, column=1, sticky='e', padx=(10, 180), pady=14)
        self.connection_badge = ctk.CTkLabel(
            hdr, text='Disconnected',
            corner_radius=10,
            fg_color='#7F1D1D', text_color='#FEE2E2', width=154)
        self.connection_badge.grid(row=0, column=1, sticky='e',
                                   padx=(10, 18), pady=14)

        # Stats row
        stats = ctk.CTkFrame(right, fg_color='transparent')
        stats.grid(row=1, column=0, sticky='ew', pady=(0, 10))
        stats.grid_columnconfigure((0, 1, 2), weight=1)
        self.port_stat = self._create_stat_card(stats, 0, 'Port', '-')
        self.tx_stat   = self._create_stat_card(stats, 1, 'TX', '0')
        self.rx_stat   = self._create_stat_card(stats, 2, 'RX', '0')

        # Offload Status Card
        offload_card = ctk.CTkFrame(right, corner_radius=14,
                                    fg_color=(CARD_LIGHT, CARD_DARK))
        offload_card.grid(row=2, column=0, sticky='ew', pady=(0, 10))
        offload_card.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(offload_card, text='Data Offload Status',
                     font=ctk.CTkFont(size=14, weight='bold')).grid(
            row=0, column=0, sticky='w', padx=16, pady=(10, 4))

        offload_metrics = ctk.CTkFrame(offload_card, fg_color='transparent')
        offload_metrics.grid(row=1, column=0, sticky='ew', padx=16, pady=(0, 8))
        offload_metrics.grid_columnconfigure(0, weight=1)
        offload_metrics.grid_columnconfigure(1, weight=1)

        left_metrics = ctk.CTkFrame(offload_metrics, fg_color='transparent')
        left_metrics.grid(row=0, column=0, sticky='nw', padx=(0, 12))

        right_metrics = ctk.CTkFrame(offload_metrics, fg_color='transparent')
        right_metrics.grid(row=0, column=1, sticky='ne')

        self.offload_status_label = ctk.CTkLabel(
            left_metrics, text='Idle',
            font=ctk.CTkFont(size=15, weight='bold'),
            text_color=('#94A3B8', '#94A3B8'))
        self.offload_status_label.grid(row=0, column=0, sticky='w', pady=(0, 3))
        
        self.offload_elapsed_label = ctk.CTkLabel(
            left_metrics, text='Elapsed: —',
            font=ctk.CTkFont(size=13),
            text_color=('#64748B', '#9CA3AF'))
        self.offload_elapsed_label.grid(row=1, column=0, sticky='w', pady=(0, 2))
        
        self.offload_duration_label = ctk.CTkLabel(
            left_metrics, text='Transfer: —',
            font=ctk.CTkFont(size=13),
            text_color=('#64748B', '#9CA3AF'))
        self.offload_duration_label.grid(row=2, column=0, sticky='w')
        
        self.offload_events_label = ctk.CTkLabel(
            right_metrics, text='Events: 0',
            font=ctk.CTkFont(size=13),
            anchor='e',
            text_color=('#64748B', '#9CA3AF'))
        self.offload_events_label.grid(row=0, column=0, sticky='e', pady=(0, 3))
        
        self.offload_connection_label = ctk.CTkLabel(
            right_metrics, text='Connection: —',
            font=ctk.CTkFont(size=13),
            anchor='e',
            text_color=('#64748B', '#9CA3AF'))
        self.offload_connection_label.grid(row=1, column=0, sticky='e', pady=(0, 2))
        
        self.offload_message_label = ctk.CTkLabel(
            right_metrics, text='Last: —',
            font=ctk.CTkFont(size=12),
            text_color=('#64748B', '#9CA3AF'),
            wraplength=280, justify='right', anchor='e')
        self.offload_message_label.grid(row=2, column=0, sticky='e')

        # Log card
        log_card = ctk.CTkFrame(right, corner_radius=14,
                                fg_color=(CARD_LIGHT, CARD_DARK))
        log_card.grid(row=3, column=0, sticky='nsew')
        log_card.grid_columnconfigure(0, weight=1)
        log_card.grid_rowconfigure(2, weight=1)

        log_hdr = ctk.CTkFrame(log_card, fg_color='transparent')
        log_hdr.grid(row=0, column=0, sticky='ew', padx=16, pady=(14, 0))
        log_hdr.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(log_hdr, text='Session Log',
                     font=ctk.CTkFont(size=16, weight='bold')).grid(
            row=0, column=0, sticky='w')

        search_frame = ctk.CTkFrame(log_hdr, fg_color=('#E5E7EB', '#374151'),
                                    corner_radius=8)
        search_frame.grid(row=0, column=1, sticky='e')
        search_frame.grid_columnconfigure(0, weight=1)
        self.search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.search_var,
            border_width=0,
            fg_color=('#F8FAFC', '#374151'),
            text_color=('#0F172A', '#F3F4F6'),
            font=ctk.CTkFont(size=14),
            width=200)
        self.search_entry.grid(row=0, column=0, padx=(8, 4), pady=4,
                               sticky='ew')
        self.search_var.trace_add('write', lambda *_: self._on_search_change())
        self.search_entry.bind('<FocusIn>',
                               lambda _e: self._on_search_focus_in())
        self.search_entry.bind('<FocusOut>',
                               lambda _e: self._on_search_focus_out())
        ctk.CTkButton(
            search_frame, text='X', width=28, height=28,
            fg_color='transparent', hover_color=('#D1D5DB', '#4B5563'),
            command=self._clear_search).grid(row=0, column=1, padx=(0, 4))
        self._activate_search_placeholder()

        legend = ctk.CTkFrame(log_card, fg_color='transparent')
        legend.grid(row=1, column=0, sticky='w', padx=16, pady=(6, 4))
        for lbl, col in [('TX', LOG_COLORS['tx']), ('RX', LOG_COLORS['rx']),
                          ('Data', LOG_COLORS['data']),
                          ('Status', LOG_COLORS['status'])]:
            ctk.CTkLabel(legend, text='*', text_color=col, width=14,
                         font=ctk.CTkFont(size=9)).pack(side='left')
            ctk.CTkLabel(legend, text=lbl, text_color=('#475569', '#94A3B8'),
                         font=ctk.CTkFont(size=11)).pack(side='left',
                                                         padx=(0, 12))

        self.log_text = ctk.CTkTextbox(log_card, wrap='none', corner_radius=10,
                                       font=('Consolas', 12))
        self.log_text.grid(row=2, column=0, sticky='nsew', padx=16,
                           pady=(0, 14))
        inner: tk.Text = self.log_text._textbox  # type: ignore[attr-defined]
        for tag, colour in LOG_COLORS.items():
            inner.tag_configure(tag, foreground=colour)
        inner.tag_configure('search_hl', background=WABASH_BLUE,
                            foreground='white')

        # Bottom bar: custom command + log tools
        bottom = ctk.CTkFrame(right, corner_radius=14,
                              fg_color=(CARD_LIGHT, CARD_DARK))
        bottom.grid(row=4, column=0, sticky='ew', pady=(10, 0))
        bottom.grid_columnconfigure(0, weight=1)

        cmd_r = ctk.CTkFrame(bottom, fg_color='transparent')
        cmd_r.grid(row=0, column=0, sticky='ew', padx=14, pady=(14, 8))
        cmd_r.grid_columnconfigure(0, weight=1)
        self.command_entry = ctk.CTkEntry(
            cmd_r, textvariable=self.command_var,
            placeholder_text='Type custom command and press Send')
        self.command_entry.grid(row=0, column=0, sticky='ew')
        ctk.CTkButton(cmd_r, text='Send', width=110,
                      fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER,
                      command=self._send_custom).grid(row=0, column=1,
                                                      padx=(10, 0))

        log_btns = ctk.CTkFrame(bottom, fg_color='transparent')
        log_btns.grid(row=1, column=0, sticky='ew', padx=14, pady=(0, 14))
        log_btns.grid_columnconfigure((0, 1), weight=1)
        ctk.CTkButton(log_btns, text='Clear Log',
                      fg_color=BTN_GREY, hover_color='#64748B',
                      command=self._clear_log).grid(
            row=0, column=0, padx=(0, 6), sticky='ew')
        ctk.CTkButton(log_btns, text='Export Log',
                      fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER,
                      command=self._export_log).grid(
            row=0, column=1, padx=(6, 0), sticky='ew')

    def _build_data_viewer_page(self) -> None:
        page = ctk.CTkFrame(self.page_container, corner_radius=0, fg_color="transparent")
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(3, weight=0)
        page.grid_rowconfigure(4, weight=1)
        self.pages["Data Viewer"] = page

        header = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            header,
            text="Data Viewer",
            font=ctk.CTkFont(size=22, weight="bold"),
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(14, 4))
        ctk.CTkLabel(
            header,
            text="Spreadsheet-style preview of saved offload workbooks",
            text_color=("#475569", "#94A3B8"),
            font=ctk.CTkFont(size=13),
        ).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 14))

        info_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        info_card.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        info_card.grid_columnconfigure(0, weight=1)
        self.viewer_path_label = ctk.CTkLabel(
            info_card,
            text="Selected Offload Workbook",
            font=ctk.CTkFont(size=14, weight="bold"),
            justify="left",
            anchor="w",
            wraplength=760,
        )
        self.viewer_path_label.grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 4))
        self.viewer_meta_label = ctk.CTkLabel(
            info_card,
            text="Truck metadata will appear here.",
            text_color=("#DCE7F7", "#E2E8F0"),
            font=ctk.CTkFont(size=18, weight="bold"),
            justify="left",
            anchor="w",
            wraplength=760,
        )
        self.viewer_meta_label.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))
        self.viewer_sheet_info_label = ctk.CTkLabel(
            info_card,
            text="Offload Rows: 0 | Total Columns: 0",
            text_color=("#93C5FD", "#93C5FD"),
            font=ctk.CTkFont(size=15, weight="bold"),
            justify="left",
            anchor="w",
        )
        self.viewer_sheet_info_label.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 12))

        controls_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        controls_card.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        controls_card.grid_columnconfigure(0, weight=1)
        controls_card.grid_columnconfigure(1, weight=0)
        controls_card.grid_columnconfigure(2, weight=0)
        controls_card.grid_columnconfigure(3, weight=0)

        self.viewer_source_label = ctk.CTkLabel(
            controls_card,
            text="Workbook: None selected",
            justify="left",
            anchor="w",
            corner_radius=8,
            fg_color=("#E5E7EB", "#374151"),
            text_color=("#0F172A", "#F3F4F6"),
            font=ctk.CTkFont(size=13, weight="bold"),
            padx=10,
            pady=8,
        )
        self.viewer_source_label.grid(row=0, column=0, sticky="ew", padx=(16, 8), pady=(12, 10))

        ctk.CTkButton(
            controls_card,
            text="Open Workbook",
            fg_color=WABASH_BLUE,
            hover_color=WABASH_BLUE_HOVER,
            command=self._open_viewer_workbook_picker,
            width=140,
        ).grid(row=0, column=1, sticky="ew", padx=8, pady=(12, 10))

        ctk.CTkButton(
            controls_card,
            text="Close Workbook",
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=self._clear_viewer_selection,
            width=130,
        ).grid(row=0, column=2, sticky="ew", padx=8, pady=(12, 10))

        self.viewer_col_window_label = ctk.CTkLabel(
            controls_card,
            text="Columns: -",
            text_color=("#475569", "#94A3B8"),
            justify="left",
            anchor="w",
        )
        self.viewer_col_window_label.grid(row=0, column=3, sticky="e", padx=(8, 16), pady=(12, 10))

        self.viewer_hint_label = ctk.CTkLabel(
            controls_card,
            text="Pinned columns stay visible. Drag the horizontal scrollbar or use arrow keys to move through the remaining data. Click any header to sort.",
            text_color=("#475569", "#94A3B8"),
            justify="left",
            anchor="w",
        )
        self.viewer_hint_label.grid(row=1, column=0, columnspan=4, sticky="ew", padx=(16, 16), pady=(0, 2))

        self.viewer_status_label = ctk.CTkLabel(
            controls_card,
            text="Open a workbook to inspect combined offload data.",
            text_color=("#475569", "#94A3B8"),
            justify="left",
            anchor="w",
        )
        self.viewer_status_label.grid(row=2, column=0, columnspan=4, sticky="ew", padx=(16, 16), pady=(0, 10))

        # --- Stats Strip (row 3) ---
        stats_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        stats_card.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        self.viewer_stats_card = stats_card
        # (name, clickable_jump_target)
        kpi_defs = [
            ("Date Range", False),
            ("Events", False),
            ("Avg Temp (°F)", False),
            ("Avg RH (%)", False),
            ("Max |Strain|", True),
            ("Max Accel Mag", True),
        ]
        for col_i, (kpi_name, clickable) in enumerate(kpi_defs):
            stats_card.grid_columnconfigure(col_i, weight=1)
            cell = ctk.CTkFrame(stats_card, corner_radius=10, fg_color=("#E5E7EB", "#1E293B"))
            cell.grid(row=0, column=col_i, padx=10, pady=10, sticky="ew")
            cell.grid_columnconfigure(0, weight=1)
            hint = " ↗ jump" if clickable else ""
            title_lbl = ctk.CTkLabel(cell, text=kpi_name + hint, font=ctk.CTkFont(size=11),
                                     text_color=("#2563EB", "#60A5FA") if clickable else ("#475569", "#94A3B8"))
            title_lbl.grid(row=0, column=0, pady=(8, 0))
            value_font_size = 15 if kpi_name == "Date Range" else 18
            val_lbl = ctk.CTkLabel(cell, text="—", font=ctk.CTkFont(size=value_font_size, weight="bold"),
                                   text_color=("#0F172A", "#F1F5F9"))
            val_lbl.grid(row=1, column=0, pady=(0, 8))
            self.viewer_kpi_labels[kpi_name] = val_lbl
            if clickable:
                for widget in (cell, title_lbl, val_lbl):
                    widget.configure(cursor="hand2")
                    widget.bind("<Button-1>", lambda _e, k=kpi_name: self._jump_to_viewer_max(k))
        stats_card.grid_remove()

        grid_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        grid_card.grid(row=4, column=0, sticky="nsew")
        grid_card.grid_columnconfigure(0, weight=1)
        grid_card.grid_rowconfigure(0, weight=1)
        self.viewer_grid_card = grid_card

        sheet_host = tk.Frame(grid_card, bd=0, highlightthickness=0, bg="#FFFFFF")
        sheet_host.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        sheet_host.grid_columnconfigure(0, weight=1)
        sheet_host.grid_rowconfigure(0, weight=1)

        self.viewer_sheet = Sheet(
            sheet_host,
            headers=[],
            data=[],
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            theme="light blue",
        )
        self.viewer_sheet.grid(row=0, column=0, sticky="nsew")
        self.viewer_sheet.enable_bindings((
            "single_select",
            "row_select",
            "column_select",
            "drag_select",
            "arrowkeys",
            "column_width_resize",
            "double_click_column_resize",
            "right_click_popup_menu",
            "rc_select",
            "copy",
            "select_all",
        ))
        self.viewer_sheet.extra_bindings("cell_select", lambda _ev: self._sync_viewer_column_label())
        if hasattr(self.viewer_sheet, "MT"):
            self.viewer_sheet.MT.bind("<Double-1>", self._on_viewer_row_double_click)

        grid_card.grid_remove()

        self._set_viewer_status("Open a workbook from the AppData data folder to start browsing offload workbooks.")

    def _set_viewer_status(self, message: str) -> None:
        if self.viewer_status_label is not None:
            self.viewer_status_label.configure(text=message)

    def _set_viewer_grid_visible(self, visible: bool) -> None:
        if self.viewer_grid_card is None:
            return
        if visible:
            self.viewer_grid_card.grid()
        else:
            self.viewer_grid_card.grid_remove()

    def _set_viewer_stats_visible(self, visible: bool) -> None:
        if self.viewer_stats_card is None:
            return
        if visible:
            self.viewer_stats_card.grid()
        else:
            self.viewer_stats_card.grid_remove()

    def _viewer_pinned_indices(self, total_cols: int) -> list[int]:
        preferred = ["Event #", "Time Stamp", "Temp", "RH"]
        pinned: list[int] = []
        for name in preferred:
            for idx, header in enumerate(self._viewer_header_row[:total_cols]):
                if idx in pinned:
                    continue
                if header.strip().lower() == name.lower():
                    pinned.append(idx)
                    break
        return pinned

    def _viewer_scrollable_indices(self, total_cols: int) -> list[int]:
        pinned = set(self._viewer_pinned_indices(total_cols))
        return [idx for idx in range(total_cols) if idx not in pinned]

    def _viewer_visible_indices(self, total_cols: int) -> list[int]:
        pinned = self._viewer_pinned_indices(total_cols)
        scrollable = self._viewer_scrollable_indices(total_cols)
        if not scrollable:
            return pinned
        max_start = max(0, len(scrollable) - self._viewer_col_span)
        self._viewer_col_start = min(max(self._viewer_col_start, 0), max_start)
        return pinned + scrollable[self._viewer_col_start:self._viewer_col_start + self._viewer_col_span]

    def _viewer_sort_key(self, value: str, column_index: int) -> tuple[int, object]:
        text = value.strip()
        if not text:
            return (2, "")
        try:
            return (0, float(text))
        except ValueError:
            pass

        header = self._viewer_header_row[column_index].strip().lower() if column_index < len(self._viewer_header_row) else ""
        if "time" in header:
            try:
                clean = text.replace(" EST", "").replace(" CST", "").replace(" MST", "").replace(" PST", "")
                return (1, datetime.datetime.strptime(clean, "%Y-%m-%d %H:%M:%S"))
            except ValueError:
                return (1, text.lower())

        return (1, text.lower())

    def _draw_viewer_range_graph(
        self,
        parent: ctk.CTkFrame,
        rows: list[list[str]],
        column_indices: list[int],
        title: str,
        x_label_mode: str = "compact",
    ) -> None:
        panel = ctk.CTkFrame(parent, corner_radius=10, fg_color=("#F8FAFC", "#0B1220"))
        panel.pack(fill="both", expand=True, padx=10, pady=10)

        header_row = ctk.CTkFrame(panel, fg_color="transparent")
        header_row.pack(fill="x", padx=12, pady=(10, 4))
        header_row.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header_row, text=title, font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, sticky="w")

        clear_pinned_button = ctk.CTkButton(
            header_row,
            text="Clear Pinned",
            width=110,
            height=28,
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
            command=lambda: None,
        )
        clear_pinned_button.grid(row=0, column=1, sticky="e")

        canvas = tk.Canvas(panel, height=230, bg="#FFFFFF", highlightthickness=0)
        canvas.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        # Build statistics per channel (column): low / average / high across the event.
        channel_stats: list[tuple[int, str, float, float, float]] = []
        all_points: list[tuple[float, int, int]] = []  # value, row_idx, col_idx
        for col_idx in column_indices:
            values: list[tuple[float, int]] = []
            for row_idx, row in enumerate(rows):
                if col_idx >= len(row):
                    continue
                try:
                    v = float(row[col_idx])
                except ValueError:
                    continue
                values.append((v, row_idx))
                all_points.append((v, row_idx, col_idx))
            if values:
                only_vals = [v for v, _ in values]
                header = self._viewer_header_row[col_idx].strip() if col_idx < len(self._viewer_header_row) else f"Col {col_idx + 1}"
                channel_stats.append((col_idx, header, min(only_vals), sum(only_vals) / len(only_vals), max(only_vals)))

        if not channel_stats:
            ctk.CTkLabel(panel, text="No numeric data available for this event.", text_color=("#64748B", "#94A3B8")).pack(anchor="w", padx=12, pady=(0, 10))
            return

        # Stats labels
        max_val, max_row, max_col = max(all_points, key=lambda t: t[0])
        min_val, min_row, min_col = min(all_points, key=lambda t: t[0])
        max_name = self._viewer_header_row[max_col] if max_col < len(self._viewer_header_row) else f"Col {max_col + 1}"
        min_name = self._viewer_header_row[min_col] if min_col < len(self._viewer_header_row) else f"Col {min_col + 1}"
        ctk.CTkLabel(
            panel,
            text=(
                f"High: {max_val:.4f} ({max_name}, sample {max_row + 1})   "
                f"Low: {min_val:.4f} ({min_name}, sample {min_row + 1})"
            ),
            text_color=("#334155", "#94A3B8"),
            font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(anchor="w", padx=12, pady=(0, 8))

        def _draw(_event: object | None = None) -> None:
            canvas.delete("all")
            width = max(canvas.winfo_width(), 320)
            height = max(canvas.winfo_height(), 230)
            left, right = 56, width - 20
            top, bottom = 16, height - 50
            canvas.create_rectangle(left, top, right, bottom, outline="#CBD5E1", width=1)

            # Symmetric axis around 0 so the zero baseline is true center reference.
            # Range = +/- (max absolute value * 1.05)
            raw_min = min(lo for _, _, lo, _, _ in channel_stats)
            raw_max = max(hi for _, _, _, _, hi in channel_stats)
            max_abs = max(abs(raw_min), abs(raw_max))
            if max_abs <= 1e-9:
                max_abs = 1.0
            y_extent = max_abs * 1.05
            y_min = -y_extent
            y_max = y_extent

            def x_for(i: int) -> float:
                if len(channel_stats) == 1:
                    return (left + right) / 2
                return left + (right - left) * (i / (len(channel_stats) - 1))

            def y_for(v: float) -> float:
                return bottom - ((v - y_min) / (y_max - y_min)) * (bottom - top)

            # Draw denser Y-axis ticks + grid for better readability.
            # Odd tick count keeps 0.00 as the middle labeled tick.
            # Doubled again from 13 to 25 for even higher vertical resolution.
            tick_count = 25
            for ti in range(tick_count):
                frac = ti / (tick_count - 1)
                y_tick = bottom - frac * (bottom - top)
                tick_val = y_min + frac * (y_max - y_min)
                if abs(tick_val) < 1e-12:
                    tick_val = 0.0
                canvas.create_line(left, y_tick, right, y_tick, fill="#E2E8F0", width=1)
                is_zero_tick = abs(tick_val) < 1e-12
                canvas.create_text(
                    10,
                    y_tick,
                    text=f"{tick_val:.4f}",
                    anchor="w",
                    fill="#111111" if is_zero_tick else "#475569",
                    font=("Calibri", 9, "bold"),
                )

            # Draw solid zero baseline like a standard XY graph.
            zero_y = y_for(0.0)
            if top <= zero_y <= bottom:
                canvas.create_line(left, zero_y, right, zero_y, fill="#111111", width=2)

            hi_coords: list[float] = []
            lo_coords: list[float] = []
            avg_coords: list[float] = []
            point_specs: list[tuple[float, float, str, str]] = []
            for i, (_, _name, lo, avg, hi) in enumerate(channel_stats):
                x = x_for(i)
                y_lo = y_for(lo)
                y_hi = y_for(hi)
                y_avg = y_for(avg)
                # vertical whisker low-high per channel
                canvas.create_line(x, y_lo, x, y_hi, fill="#94A3B8", width=1)
                hi_coords.extend([x, y_hi])
                lo_coords.extend([x, y_lo])
                avg_coords.extend([x, y_avg])

            if len(hi_coords) >= 4:
                canvas.create_line(*hi_coords, fill="#DC2626", width=2)
            if len(lo_coords) >= 4:
                canvas.create_line(*lo_coords, fill="#2563EB", width=2)
            if len(avg_coords) >= 4:
                canvas.create_line(*avg_coords, fill="#0F766E", width=2)

            # X labels: compact channel names (S# / Ax#...) or just sensor number.
            def _compact_channel(name: str) -> str:
                n = name.strip().lower().replace(" ", "")
                if n.startswith("strain"):
                    suffix = "1"
                    if "_" in n:
                        tail = n.rsplit("_", 1)[-1]
                        if tail.isdigit():
                            suffix = tail
                    return f"S{suffix}"
                if n.startswith("accelx"):
                    suffix = "1"
                    if "_" in n:
                        tail = n.rsplit("_", 1)[-1]
                        if tail.isdigit():
                            suffix = tail
                    return f"Ax{suffix}"
                if n.startswith("accely"):
                    suffix = "1"
                    if "_" in n:
                        tail = n.rsplit("_", 1)[-1]
                        if tail.isdigit():
                            suffix = tail
                    return f"Ay{suffix}"
                if n.startswith("accelz"):
                    suffix = "1"
                    if "_" in n:
                        tail = n.rsplit("_", 1)[-1]
                        if tail.isdigit():
                            suffix = tail
                    return f"Az{suffix}"
                return name.replace(" ", "")

            def _number_only(name: str) -> str:
                n = name.strip().lower().replace(" ", "")
                if "_" in n:
                    tail = n.rsplit("_", 1)[-1]
                    if tail.isdigit():
                        return tail
                return "1"

            for i, (_col_idx, name, _lo, _avg, _hi) in enumerate(channel_stats):
                x = x_for(i)
                label = _number_only(name) if x_label_mode == "number" else _compact_channel(name)
                canvas.create_text(x, height - 6, text=label, anchor="s", fill="#475569", font=("Calibri", 9))

            # Collect interactive scatter points for exact coordinate lookup.
            for i, (_col_idx, name, lo, avg, hi) in enumerate(channel_stats):
                x = x_for(i)
                x_label = _number_only(name) if x_label_mode == "number" else _compact_channel(name)
                point_specs.append((x, y_for(hi), f"x={x_label}, y={hi:.6f}", "#DC2626"))
                point_specs.append((x, y_for(avg), f"x={x_label}, y={avg:.6f}", "#0F766E"))
                point_specs.append((x, y_for(lo), f"x={x_label}, y={lo:.6f}", "#2563EB"))

            hover_tooltip_ids: list[int] = []
            pinned_tooltips: dict[tuple[float, float, str], list[int]] = {}

            def _find_nearest_point(mx: float, my: float, max_px: float = 10.0) -> tuple[float, float, str] | None:
                best: tuple[float, float, str] | None = None
                best_d2 = max_px * max_px
                for px, py, ptext, _color in point_specs:
                    d2 = (mx - px) * (mx - px) + (my - py) * (my - py)
                    if d2 <= best_d2:
                        best = (px, py, ptext)
                        best_d2 = d2
                return best

            def _hide_hover_tooltip() -> None:
                for item in list(hover_tooltip_ids):
                    try:
                        canvas.delete(item)
                    except tk.TclError:
                        pass
                hover_tooltip_ids.clear()

            def _draw_tooltip(px: float, py: float, text: str, fill: str = "#FEF9C3", outline: str = "#F59E0B") -> list[int]:
                try:
                    tx = min(max(px + 8, left + 4), right - 4)
                    ty = max(py - 12, top + 4)
                    text_id = canvas.create_text(tx, ty, text=text, anchor="sw", fill="#0F172A", font=("Calibri", 10, "bold"))
                    bbox = canvas.bbox(text_id)
                    if bbox is None:
                        return []
                    x1, y1, x2, y2 = bbox
                    bg_id = canvas.create_rectangle(x1 - 6, y1 - 4, x2 + 6, y2 + 4, fill=fill, outline=outline, width=1)
                    canvas.tag_raise(text_id, bg_id)
                    return [bg_id, text_id]
                except tk.TclError:
                    return []

            for px, py, _ptext, color in point_specs:
                canvas.create_oval(px - 3, py - 3, px + 3, py + 3, fill=color, outline="", width=0)

            def _on_motion(event: object) -> None:
                try:
                    ex = float(getattr(event, "x", 0))
                    ey = float(getattr(event, "y", 0))
                    nearest = _find_nearest_point(ex, ey)
                    if nearest is None:
                        _hide_hover_tooltip()
                        return
                    px, py, ptext = nearest
                    _hide_hover_tooltip()
                    hover_tooltip_ids.extend(_draw_tooltip(px, py, ptext))
                except Exception:
                    pass

            def _on_click(event: object) -> str:
                try:
                    ex = float(getattr(event, "x", 0))
                    ey = float(getattr(event, "y", 0))
                    nearest = _find_nearest_point(ex, ey)
                    if nearest is None:
                        _hide_hover_tooltip()
                        return "break"
                    px, py, ptext = nearest
                    key = (px, py, ptext)
                    if key in pinned_tooltips:
                        for item in pinned_tooltips[key]:
                            try:
                                canvas.delete(item)
                            except tk.TclError:
                                pass
                        pinned_tooltips.pop(key, None)
                    else:
                        ids = _draw_tooltip(px, py, ptext, fill="#DBEAFE", outline="#1D4ED8")
                        if ids:
                            pinned_tooltips[key] = ids
                    return "break"
                except Exception:
                    pass
                    return "break"

            def _on_leave(_event: object) -> None:
                try:
                    _hide_hover_tooltip()
                except Exception:
                    pass

            def _clear_all_pins() -> None:
                try:
                    for ids in list(pinned_tooltips.values()):
                        for item in ids:
                            try:
                                canvas.delete(item)
                            except tk.TclError:
                                pass
                    pinned_tooltips.clear()
                    _hide_hover_tooltip()
                except Exception:
                    pass

            clear_pinned_button.configure(command=_clear_all_pins)

            canvas.bind("<Motion>", _on_motion)
            canvas.bind("<Button-1>", _on_click)
            canvas.bind("<Leave>", _on_leave)

        canvas.bind("<Configure>", _draw)
        _draw()

    def _show_viewer_event_detail(self, row_index: int) -> None:
        rows = self._viewer_filtered_rows if self._viewer_filtered_rows else self._viewer_sheet_rows
        if row_index < 0 or row_index >= len(rows):
            return

        event_col = next((i for i, h in enumerate(self._viewer_header_row) if h.strip().lower() == "event #"), None)
        time_col = next((i for i, h in enumerate(self._viewer_header_row) if h.strip().lower() == "time stamp"), None)
        strain_cols = [i for i, h in enumerate(self._viewer_header_row) if h.strip().lower() == "strain" or h.strip().lower().startswith("strain_")]
        accel_cols = [i for i, h in enumerate(self._viewer_header_row) if h.strip().lower().startswith("accel")]

        def _accel_axis_groups() -> dict[str, list[int]]:
            groups: dict[str, list[tuple[int, int]]] = {"x": [], "y": [], "z": []}
            for col_idx in accel_cols:
                if col_idx >= len(self._viewer_header_row):
                    continue
                header = self._viewer_header_row[col_idx].strip().lower().replace(" ", "")
                axis = ""
                if "accelx" in header or "accel_x" in header:
                    axis = "x"
                elif "accely" in header or "accel_y" in header:
                    axis = "y"
                elif "accelz" in header or "accel_z" in header:
                    axis = "z"
                if not axis:
                    continue
                sensor_num = 1
                if "_" in header:
                    tail = header.rsplit("_", 1)[-1]
                    if tail.isdigit():
                        sensor_num = int(tail)
                groups[axis].append((sensor_num, col_idx))
            return {axis: [c for _, c in sorted(vals, key=lambda t: t[0])] for axis, vals in groups.items()}

        selected_row = rows[row_index]
        selected_event = None
        if event_col is not None and event_col < len(selected_row):
            raw = selected_row[event_col].strip()
            if raw:
                try:
                    selected_event = int(float(raw))
                except ValueError:
                    selected_event = raw

        if selected_event is None:
            event_rows = [selected_row]
            event_title = "Selected Sample"
        else:
            event_rows = []
            for row in rows:
                if event_col is None or event_col >= len(row):
                    continue
                raw = row[event_col].strip()
                try:
                    value: object = int(float(raw))
                except ValueError:
                    value = raw
                if value == selected_event:
                    event_rows.append(row)
            event_title = f"Event # {selected_event}"

        start_time = "-"
        end_time = "-"
        if time_col is not None and event_rows:
            stamps = [r[time_col] for r in event_rows if time_col < len(r) and r[time_col].strip() and r[time_col].strip().lower() != "time not set"]
            if stamps:
                start_time = stamps[0]
                end_time = stamps[-1]

        if self._viewer_detail_window is not None:
            try:
                self._viewer_detail_window.destroy()
            except Exception:
                pass

        win = ctk.CTkToplevel(self.root)
        win.title(f"{event_title} Analysis")
        win.geometry("980x700")
        win.minsize(860, 620)
        win.grab_set()
        self._viewer_detail_window = win

        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(win, text=f"{event_title} Analysis", font=ctk.CTkFont(size=18, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 4)
        )
        ctk.CTkLabel(
            win,
            text=f"Samples: {len(event_rows)} | Time: {start_time} -> {end_time}",
            text_color=("#475569", "#94A3B8"),
            font=ctk.CTkFont(size=12, weight="bold"),
        ).grid(row=0, column=0, sticky="e", padx=16, pady=(14, 4))

        tabs = ctk.CTkTabview(win)
        tabs.grid(row=1, column=0, sticky="nsew", padx=12, pady=(8, 12))

        tab_strain = tabs.add("Strain")
        tab_accel = tabs.add("Acceleration")
        tab_dump = tabs.add("Data Dump")

        if strain_cols:
            self._draw_viewer_range_graph(tab_strain, event_rows, strain_cols, "Strain Envelope by Channel", x_label_mode="number")
        else:
            ctk.CTkLabel(tab_strain, text="No strain columns in this workbook.").pack(anchor="w", padx=10, pady=10)

        if accel_cols:
            axis_groups = _accel_axis_groups()
            accel_tabs = ctk.CTkTabview(tab_accel)
            accel_tabs.pack(fill="both", expand=True, padx=6, pady=0)
            tab_x = accel_tabs.add("X")
            tab_y = accel_tabs.add("Y")
            tab_z = accel_tabs.add("Z")

            if axis_groups["x"]:
                self._draw_viewer_range_graph(tab_x, event_rows, axis_groups["x"], "Acceleration X by Sensor", x_label_mode="number")
            else:
                ctk.CTkLabel(tab_x, text="No Accel X columns in this workbook.").pack(anchor="w", padx=10, pady=10)

            if axis_groups["y"]:
                self._draw_viewer_range_graph(tab_y, event_rows, axis_groups["y"], "Acceleration Y by Sensor", x_label_mode="number")
            else:
                ctk.CTkLabel(tab_y, text="No Accel Y columns in this workbook.").pack(anchor="w", padx=10, pady=10)

            if axis_groups["z"]:
                self._draw_viewer_range_graph(tab_z, event_rows, axis_groups["z"], "Acceleration Z by Sensor", x_label_mode="number")
            else:
                ctk.CTkLabel(tab_z, text="No Accel Z columns in this workbook.").pack(anchor="w", padx=10, pady=10)
        else:
            ctk.CTkLabel(tab_accel, text="No acceleration columns in this workbook.").pack(anchor="w", padx=10, pady=10)

        dump = ctk.CTkTextbox(tab_dump, wrap="none")
        dump.pack(fill="both", expand=True, padx=10, pady=10)
        headers = [h.strip() if h.strip() else f"Column {i + 1}" for i, h in enumerate(self._viewer_header_row)]
        lines = [",".join(headers)]
        for row in event_rows:
            normalized = [row[i] if i < len(row) else "" for i in range(len(headers))]
            lines.append(",".join(normalized))
        dump.insert("1.0", "\n".join(lines))
        dump.configure(state="disabled")

    def _on_viewer_row_double_click(self, event: object) -> None:
        if self.viewer_sheet is None:
            return
        selected = self.viewer_sheet.get_currently_selected()
        if not selected:
            return
        rows = self._viewer_filtered_rows or self._viewer_sheet_rows
        if not rows:
            return
        row_idx = getattr(selected, "row", None)
        if row_idx is None and isinstance(selected, tuple) and len(selected) >= 1:
            row_idx = selected[0]
        if row_idx is None or not isinstance(row_idx, int) or row_idx < 0 or row_idx >= len(rows):
            return
        self._show_viewer_event_detail(row_idx)

    def _on_viewer_header_click(self, column_index: int) -> None:
        if column_index < 0:
            return
        if self._viewer_sort_column == column_index:
            self._viewer_sort_descending = not self._viewer_sort_descending
        else:
            self._viewer_sort_column = column_index
            self._viewer_sort_descending = False
        self._viewer_sheet_rows.sort(
            key=lambda row: self._viewer_sort_key(row[column_index] if column_index < len(row) else "", column_index),
            reverse=self._viewer_sort_descending,
        )
        self._apply_viewer_filter()

    def _viewer_default_open_dir(self) -> Path:
        if self._viewer_workbook_path is not None:
            return self._viewer_workbook_path.parent
        return _default_data_path()

    def _refresh_data_viewer(self) -> None:
        if self.viewer_source_label is not None:
            if self._viewer_workbook_path is not None:
                self.viewer_source_label.configure(text=f"Workbook: {self._viewer_workbook_path}")
            else:
                self.viewer_source_label.configure(text="Workbook: None selected")
        if self._viewer_workbook_path is None:
            self._set_viewer_status("Open a workbook from the AppData data folder to start browsing offload workbooks.")
            self._update_viewer_column_window_label(0)
            return
        if self._viewer_workbook_path.exists():
            self._load_viewer_workbook()
        else:
            missing_name = self._viewer_workbook_path.name
            self._clear_viewer_selection()
            self._set_viewer_status(f"Previously selected workbook '{missing_name}' is no longer available.")

    def _open_viewer_workbook_picker(self) -> None:
        initial_dir = self._viewer_default_open_dir()
        try:
            initial_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        selected = filedialog.askopenfilename(
            title="Open Offload Workbook",
            initialdir=str(initial_dir),
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
        )
        if not selected:
            return

        self._viewer_workbook_path = Path(selected)
        self._load_viewer_workbook()

    def _clear_viewer_selection(self) -> None:
        self._viewer_workbook_path = None
        self._viewer_col_start = 0
        self._viewer_sort_column = None
        self._viewer_sort_descending = False
        self._viewer_strain_low_thresh = 0.0
        self._viewer_strain_high_thresh = 0.0
        self._viewer_max_strain_row_idx = -1
        self._viewer_max_accel_row_idx = -1
        self._viewer_max_strain_col_idx = -1
        self._viewer_max_accel_col_idx = -1
        self.viewer_sheet_var.set("No sheets")
        self._viewer_sheet_rows = []
        self._viewer_filtered_rows = []
        self._viewer_header_row = []
        self._populate_viewer_tree([])
        self._set_viewer_grid_visible(False)
        self._set_viewer_stats_visible(False)
        if self._viewer_detail_window is not None:
            try:
                self._viewer_detail_window.destroy()
            except Exception:
                pass
            self._viewer_detail_window = None
        if self.viewer_source_label is not None:
            self.viewer_source_label.configure(text="Workbook: None selected")
        if self.viewer_path_label is not None:
            self.viewer_path_label.configure(text="Selected Offload Workbook")
        if self.viewer_meta_label is not None:
            self.viewer_meta_label.configure(text="Truck metadata will appear here.")
        if self.viewer_sheet_info_label is not None:
            self.viewer_sheet_info_label.configure(text="Offload Rows: 0 | Total Columns: 0")
        self._update_viewer_column_window_label(0)
        self._set_viewer_status("Open a workbook to inspect combined offload data.")

    def _load_viewer_workbook(self) -> None:
        workbook_path = self._viewer_workbook_path
        if workbook_path is None:
            self.viewer_sheet_var.set("No sheets")
            self._viewer_sheet_rows = []
            self._populate_viewer_tree([])
            if self.viewer_path_label is not None:
                self.viewer_path_label.configure(text="Selected Offload Workbook")
            if self.viewer_meta_label is not None:
                self.viewer_meta_label.configure(text="Truck metadata will appear here.")
            if self.viewer_sheet_info_label is not None:
                self.viewer_sheet_info_label.configure(text="Offload Rows: 0 | Total Columns: 0")
            self._set_viewer_grid_visible(False)
            self._set_viewer_stats_visible(False)
            return

        if not workbook_path.exists():
            self._clear_viewer_selection()
            self._set_viewer_status("Selected workbook could not be found.")
            return

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            sheet_names = list(workbook.sheetnames)
            workbook.close()
        except Exception as exc:
            self._set_viewer_status(f"Could not open workbook: {exc}")
            return

        current_sheet = self.viewer_sheet_var.get().strip()
        if sheet_names:
            if current_sheet not in sheet_names:
                current_sheet = sheet_names[0]
            self.viewer_sheet_var.set(current_sheet)
            self._load_viewer_sheet_data()
        else:
            self.viewer_sheet_var.set("No sheets")
            self._viewer_sheet_rows = []
            self._populate_viewer_tree([])

    def _load_viewer_sheet_data(self) -> None:
        workbook_path = self._viewer_workbook_path
        sheet_name = self.viewer_sheet_var.get().strip()
        if workbook_path is None or not sheet_name or sheet_name == "No sheets":
            self._viewer_sheet_rows = []
            self._populate_viewer_tree([])
            return

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                self._set_viewer_status("Selected sheet no longer exists in this workbook.")
                return
            worksheet = workbook[sheet_name]
            rows: list[list[str]] = []
            max_cols = 0
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                max_cols = max(max_cols, len(values))
                rows.append(values)
            workbook.close()
        except Exception as exc:
            self._set_viewer_status(f"Could not load sheet data: {exc}")
            return

        normalized_rows = [row + [""] * (max_cols - len(row)) for row in rows]
        if len(normalized_rows) >= 4:
            self._viewer_header_row = list(normalized_rows[3])
            self._viewer_sheet_rows = normalized_rows[4:]
        else:
            self._viewer_header_row = [f"Column {idx + 1}" for idx in range(max_cols)]
            self._viewer_sheet_rows = normalized_rows
        self._viewer_col_start = 0
        self._apply_viewer_filter()
        self._set_viewer_grid_visible(bool(self._viewer_sheet_rows))

        if self.viewer_source_label is not None:
            self.viewer_source_label.configure(text=f"Workbook: {workbook_path}")
        if self.viewer_path_label is not None:
            session_name = workbook_path.parent.name
            truck_name = workbook_path.parent.parent.name if workbook_path.parent.parent != workbook_path.parent else "Unknown"
            self.viewer_path_label.configure(text=f"Offload Session: {truck_name} / {session_name}")
        truck_meta = normalized_rows[0][1] if len(normalized_rows) > 0 and len(normalized_rows[0]) > 1 else "Unknown"
        desc_meta = normalized_rows[1][1] if len(normalized_rows) > 1 and len(normalized_rows[1]) > 1 else ""
        if self.viewer_meta_label is not None:
            self.viewer_meta_label.configure(
                text=f"Truck ID: {truck_meta} | Description: {desc_meta or '-'}"
            )
        data_rows = len(self._viewer_sheet_rows)
        if self.viewer_sheet_info_label is not None:
            self.viewer_sheet_info_label.configure(text=f"Offload Rows: {data_rows} | Total Columns: {max_cols}")
        self._set_viewer_status(f"Viewing sheet '{sheet_name}' from {workbook_path.name}.")

    def _apply_viewer_filter(self) -> None:
        self._viewer_filtered_rows = list(self._viewer_sheet_rows)
        self._populate_viewer_tree(self._viewer_filtered_rows)
        self._update_viewer_stats(self._viewer_filtered_rows)
        self._set_viewer_status(f"Rows shown: {len(self._viewer_filtered_rows)}")

    def _update_viewer_stats(self, rows: list[list[str]]) -> None:
        if not self.viewer_kpi_labels:
            return
        if not rows or not self._viewer_header_row:
            for lbl in self.viewer_kpi_labels.values():
                lbl.configure(text="—")
            self._set_viewer_stats_visible(False)
            return

        def _col(name: str) -> int | None:
            return next((i for i, h in enumerate(self._viewer_header_row) if h.strip().lower() == name.lower()), None)

        event_col = _col("Event #")
        temp_col = _col("Temp")
        time_col = _col("Time Stamp")

        # Collect all strain column indices (Strain, Strain_2, Strain_3, ...)
        strain_cols = [i for i, h in enumerate(self._viewer_header_row)
                       if h.strip().lower() == "strain" or h.strip().lower().startswith("strain_")]
        # Collect all accel column indices
        accel_cols = [i for i, h in enumerate(self._viewer_header_row)
                      if h.strip().lower().startswith("accel")]

        # Events
        if event_col is not None:
            event_nums = set()
            for row in rows:
                if event_col < len(row) and row[event_col].strip():
                    try:
                        event_nums.add(int(float(row[event_col])))
                    except ValueError:
                        pass
            self.viewer_kpi_labels["Events"].configure(text=str(len(event_nums)))
        else:
            self.viewer_kpi_labels["Events"].configure(text=str(len(rows)))

        # Max |Strain| across all strain columns
        max_strain = 0.0
        self._viewer_max_strain_row_idx = -1
        self._viewer_max_strain_col_idx = -1
        for row_i, row in enumerate(rows):
            for ci in strain_cols:
                if ci < len(row):
                    try:
                        v = abs(float(row[ci]))
                        if v > max_strain:
                            max_strain = v
                            self._viewer_max_strain_row_idx = row_i
                            self._viewer_max_strain_col_idx = ci
                    except ValueError:
                        pass
        self.viewer_kpi_labels["Max |Strain|"].configure(
            text=f"{max_strain:.3f}" if strain_cols else "—"
        )

        # Max Accel value from actual accel cells (ensures KPI jump lands on a real cell)
        max_accel_mag = 0.0
        self._viewer_max_accel_row_idx = -1
        self._viewer_max_accel_col_idx = -1
        for row_i, row in enumerate(rows):
            for col_i in accel_cols:
                if col_i >= len(row):
                    continue
                try:
                    comp_abs = abs(float(row[col_i]))
                except ValueError:
                    continue
                if comp_abs > max_accel_mag:
                    max_accel_mag = comp_abs
                    self._viewer_max_accel_row_idx = row_i
                    self._viewer_max_accel_col_idx = col_i
        self.viewer_kpi_labels["Max Accel Mag"].configure(
            text=f"{max_accel_mag:.3f} g" if accel_cols else "—"
        )

        # Avg Temp
        if temp_col is not None:
            temps = []
            for row in rows:
                if temp_col < len(row):
                    try:
                        temps.append(float(row[temp_col]))
                    except ValueError:
                        pass
            self.viewer_kpi_labels["Avg Temp (°F)"].configure(
                text=f"{sum(temps)/len(temps):.1f} °F" if temps else "—"
            )
        else:
            self.viewer_kpi_labels["Avg Temp (°F)"].configure(text="—")

        # Avg RH
        rh_col = _col("RH")
        if rh_col is not None:
            rh_vals = []
            for row in rows:
                if rh_col < len(row):
                    try:
                        rh_vals.append(float(row[rh_col]))
                    except ValueError:
                        pass
            self.viewer_kpi_labels["Avg RH (%)"].configure(
                text=f"{sum(rh_vals)/len(rh_vals):.1f} %" if rh_vals else "—"
            )
        else:
            self.viewer_kpi_labels["Avg RH (%)"].configure(text="—")

        # Date Range
        if time_col is not None:
            dates = []
            for row in rows:
                if time_col < len(row):
                    val = row[time_col].strip()
                    if val and val.lower() != "time not set":
                        try:
                            clean = val.split()[0]  # just date portion
                            dates.append(clean)
                        except Exception:
                            pass
            if dates:
                dates_sorted = sorted(set(dates))
                if len(dates_sorted) == 1:
                    self.viewer_kpi_labels["Date Range"].configure(text=dates_sorted[0])
                else:
                    self.viewer_kpi_labels["Date Range"].configure(text=f"{dates_sorted[0]} -> {dates_sorted[-1]}")
            else:
                self.viewer_kpi_labels["Date Range"].configure(text="—")
        else:
            self.viewer_kpi_labels["Date Range"].configure(text="—")

        # Compute thresholds from max |strain| using fixed percentage bands:
        # 0-33% green, 33-66% yellow, 66-100% red.
        strain_magnitudes: list[float] = []
        for row in rows:
            for ci in strain_cols:
                if ci < len(row):
                    try:
                        strain_magnitudes.append(abs(float(row[ci])))
                    except ValueError:
                        pass

        if strain_magnitudes:
            max_v = max(strain_magnitudes)
            if max_v > 0:
                self._viewer_strain_low_thresh = max_v * 0.33
                self._viewer_strain_high_thresh = max_v * 0.66
            else:
                # All zero values: keep everything green.
                self._viewer_strain_low_thresh = float("inf")
                self._viewer_strain_high_thresh = float("inf")
        else:
            # No numeric strain values: keep everything green.
            self._viewer_strain_low_thresh = float("inf")
            self._viewer_strain_high_thresh = float("inf")

        self._set_viewer_stats_visible(True)

    def _jump_to_viewer_max(self, kpi_name: str) -> None:
        if self.viewer_sheet is None:
            return
        rows = self._viewer_filtered_rows if self._viewer_filtered_rows else self._viewer_sheet_rows
        if not rows:
            return

        if kpi_name == "Max |Strain|":
            target_row_idx = self._viewer_max_strain_row_idx
            target_col_idx = self._viewer_max_strain_col_idx
        else:
            target_row_idx = self._viewer_max_accel_row_idx
            target_col_idx = self._viewer_max_accel_col_idx

        if target_row_idx < 0:
            return

        if 0 <= target_col_idx < len(self._viewer_header_row):
            self.viewer_sheet.see(target_row_idx, target_col_idx, redraw=False)
            self.viewer_sheet.select_cell(target_row_idx, target_col_idx, redraw=True)
        else:
            self.viewer_sheet.see(target_row_idx, 0, redraw=False)
            self.viewer_sheet.select_cell(target_row_idx, 0, redraw=True)

        if 0 <= target_col_idx < len(self._viewer_header_row) and target_row_idx < len(rows) and target_col_idx < len(rows[target_row_idx]):
            self._set_viewer_status(
                f"Jumped to {kpi_name}: {self._viewer_header_row[target_col_idx]} = {rows[target_row_idx][target_col_idx]}"
            )

    def _populate_viewer_tree(self, rows: list[list[str]]) -> None:
        if self.viewer_sheet is None:
            return
        max_cols = max(max((len(row) for row in rows), default=0), len(self._viewer_header_row))
        if max_cols == 0:
            self.viewer_sheet.headers([])
            self.viewer_sheet.set_sheet_data([])
            self._update_viewer_column_window_label(0)
            self._set_viewer_grid_visible(False)
            return

        # tksheet displays full column set natively with true cell selection/highlighting.
        normalized = [row + [""] * (max_cols - len(row)) for row in rows]
        headers = [(self._viewer_header_row[i].strip() if i < len(self._viewer_header_row) else _excel_column_name(i)) for i in range(max_cols)]
        self.viewer_sheet.headers(headers, redraw=False)
        self.viewer_sheet.set_sheet_data(normalized, redraw=False, reset_highlights=True)

        self.viewer_sheet.redraw()
        self._update_viewer_column_window_label(max_cols)
        self._set_viewer_grid_visible(True)

    def _update_viewer_column_window_label(self, total_cols: int) -> None:
        if self.viewer_col_window_label is None:
            return
        if total_cols <= 0:
            self.viewer_col_window_label.configure(text="Columns: -")
            return
        self.viewer_col_window_label.configure(text=f"Columns: {total_cols}")

    def _sync_viewer_column_label(self) -> None:
        if self.viewer_col_window_label is None or self.viewer_sheet is None:
            return
        sel = self.viewer_sheet.get_currently_selected()
        if not sel:
            return
        col_idx = getattr(sel, "column", None)
        if col_idx is None and isinstance(sel, tuple) and len(sel) >= 2:
            col_idx = sel[1]
        if isinstance(col_idx, int) and 0 <= col_idx < len(self._viewer_header_row):
            self.viewer_col_window_label.configure(text=f"Column: {self._viewer_header_row[col_idx]}")

    def _on_viewer_xscroll(self, *args: str) -> None:
        rows = self._viewer_filtered_rows if self._viewer_filtered_rows else self._viewer_sheet_rows
        total_cols = max(max((len(row) for row in rows), default=0), len(self._viewer_header_row))
        scrollable = self._viewer_scrollable_indices(total_cols)
        if len(scrollable) <= self._viewer_col_span:
            self._viewer_col_start = 0
            self._populate_viewer_tree(rows)
            return

        max_start = max(0, len(scrollable) - self._viewer_col_span)
        if not args:
            return

        if args[0] == "moveto" and len(args) > 1:
            try:
                fraction = float(args[1])
            except ValueError:
                return
            self._viewer_col_start = int(round(fraction * max_start))
        elif args[0] == "scroll" and len(args) > 2:
            try:
                step = int(args[1])
            except ValueError:
                return
            page_size = max(1, self._viewer_col_span // 2)
            delta = step * (1 if args[2] == "units" else page_size)
            self._viewer_col_start = min(max(self._viewer_col_start + delta, 0), max_start)
        else:
            return

        self._populate_viewer_tree(rows)

    def _shift_viewer_columns(self, delta: int) -> None:
        rows = self._viewer_filtered_rows if self._viewer_filtered_rows else self._viewer_sheet_rows
        total_cols = max(max((len(row) for row in rows), default=0), len(self._viewer_header_row))
        scrollable = self._viewer_scrollable_indices(total_cols)
        if len(scrollable) <= self._viewer_col_span:
            self._viewer_col_start = 0
        else:
            self._viewer_col_start = min(max(self._viewer_col_start + delta, 0), len(scrollable) - self._viewer_col_span)
        self._populate_viewer_tree(rows)

    def _open_selected_viewer_folder(self) -> None:
        target = self._viewer_workbook_path
        path_to_open = target.parent if target is not None else self._viewer_default_open_dir()
        try:
            path_to_open.mkdir(parents=True, exist_ok=True)
            if sys.platform == "win32":
                os.startfile(str(path_to_open))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path_to_open)])
            else:
                subprocess.Popen(["xdg-open", str(path_to_open)])
        except Exception as exc:
            messagebox.showerror("Open Folder Error", str(exc))

    def _open_viewer_default_folder(self) -> None:
        path_to_open = _default_data_path()
        try:
            path_to_open.mkdir(parents=True, exist_ok=True)
            if sys.platform == "win32":
                os.startfile(str(path_to_open))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path_to_open)])
            else:
                subprocess.Popen(["xdg-open", str(path_to_open)])
        except Exception as exc:
            messagebox.showerror("Open Folder Error", str(exc))

    def _set_theme(self, mode: str) -> None:
        ctk.set_appearance_mode(mode)

    def _set_text_scale(self, value: float) -> None:
        ctk.set_widget_scaling(value)
        self.text_scale_label.configure(text=f"{int(value * 100)}%")

    def _activate_search_placeholder(self) -> None:
        self.search_placeholder_active = True
        self.search_var.set("Find...")
        self.search_entry.configure(text_color=("#64748B", "#9CA3AF"))

    def _deactivate_search_placeholder(self) -> None:
        if self.search_placeholder_active:
            self.search_placeholder_active = False
            self.search_var.set("")
            self.search_entry.configure(text_color=("#0F172A", "#F3F4F6"))

    def _effective_search_term(self) -> str:
        if self.search_placeholder_active:
            return ""
        return self.search_var.get().strip()

    def _on_search_focus_in(self) -> None:
        self._deactivate_search_placeholder()

    def _on_search_focus_out(self) -> None:
        if not self.search_var.get().strip():
            self._activate_search_placeholder()

    def _schedule_search_highlight(self) -> None:
        if self._search_highlight_job is not None:
            self.root.after_cancel(self._search_highlight_job)
        self._search_highlight_job = self.root.after(120, self._run_search_highlight)

    def _run_search_highlight(self) -> None:
        self._search_highlight_job = None
        self._highlight_search()

    def _set_density(self, mode: str) -> None:
        px, py, fs, rp = DENSITY[mode]
        self.log_text.configure(font=("Consolas", fs))

    def _highlight_search(self) -> None:
        # Guard against callbacks during initialization before log_text is created
        if not hasattr(self, 'log_text'):
            return
        
        inner: tk.Text = self.log_text._textbox  # type: ignore[attr-defined]
        inner.tag_remove("search_hl", "1.0", "end")
        term = self._effective_search_term()
        if not term:
            return
        start = "1.0"
        while True:
            pos = inner.search(term, start, stopindex="end", nocase=True)
            if not pos:
                break
            end = f"{pos}+{len(term)}c"
            inner.tag_add("search_hl", pos, end)
            start = end

    def _on_search_change(self) -> None:
        self._schedule_search_highlight()

    def _clear_search(self) -> None:
        self.search_var.set("")
        self.search_entry.focus_set()
        self._deactivate_search_placeholder()

    def _refresh_ports(self) -> None:
        ports = [port.device for port in list_ports.comports()]
        if not ports:
            ports = ["No ports found"]

        self.port_combo.configure(values=ports)

        current_port = self.port_var.get()
        if current_port not in ports:
            self.port_var.set(ports[0])

        if self.serial_service.is_connected:
            self.port_stat.configure(text=self.connected_port if self.connected_port else "-")

        self._update_port_selection_status()

    def _connect(self, suppress_warnings: bool = False) -> bool:
        port = self.port_var.get().strip()
        if not port or port == "No ports found":
            if not suppress_warnings:
                messagebox.showwarning("Port Required", "Select a COM port before connecting.")
            return False

        try:
            baud = int(self.baud_var.get().strip())
        except ValueError:
            if not suppress_warnings:
                messagebox.showerror("Invalid Baud", "Baud rate must be an integer.")
            return False

        detected_role = self.detected_port_roles.get(port)
        if detected_role is None or detected_role == "unknown":
            detected_role = self._probe_port_role(port, baud)
            self.detected_port_roles[port] = detected_role
            self._update_port_selection_status()

        if detected_role == "receiver":
            self._set_connected_state(False, "-", role="disconnected")
            if not suppress_warnings:
                messagebox.showwarning("Receiver Detected", "The selected port appears to be a receiver. Connect to the transmitter instead.")
            return False

        try:
            self.serial_service.connect(SerialConfig(port=port, baudrate=baud))
            connection_role = detected_role if detected_role in {"transmitter", "unknown"} else "unknown"
            self._set_connected_state(True, port, role=connection_role)
            return True
        except Exception as exc:
            self._set_connected_state(False, "-", role="disconnected")
            if not suppress_warnings:
                messagebox.showerror("Connection Error", str(exc))
            return False

    def _disconnect(self) -> None:
        self.serial_service.disconnect()
        self._set_connected_state(False, "-", role="disconnected")

    def _set_connected_state(self, is_connected: bool, port_text: str, role: str = "unknown") -> None:
        self.connected = is_connected
        self.connected_role = role if is_connected else "disconnected"
        self.connected_port = port_text if is_connected else "-"

        connect_fg = BTN_GREY if is_connected else WABASH_BLUE
        connect_hover = BTN_GREY_HOVER if is_connected else WABASH_BLUE_HOVER
        disconnect_fg = WABASH_BLUE if is_connected else BTN_GREY
        disconnect_hover = WABASH_BLUE_HOVER if is_connected else BTN_GREY_HOVER

        self.connect_button.configure(fg_color=connect_fg, hover_color=connect_hover)
        self.disconnect_button.configure(fg_color=disconnect_fg, hover_color=disconnect_hover)
        self.sidebar_connect_button.configure(fg_color=connect_fg, hover_color=connect_hover)
        self.sidebar_disconnect_button.configure(fg_color=disconnect_fg, hover_color=disconnect_hover)

        if is_connected:
            if role == "transmitter":
                self.connection_badge.configure(text="Connected", fg_color="#14532D", text_color="#DCFCE7")
                self.sidebar_status.configure(text="Connected", fg_color="#14532D", text_color="#DCFCE7")
                self.db_tx_status.configure(text="Connected", text_color=("#22C55E", "#22C55E"))
            else:
                self.connection_badge.configure(text="Verifying", fg_color="#92400E", text_color="#FEF3C7")
                self.sidebar_status.configure(text="Verifying", fg_color="#92400E", text_color="#FEF3C7")
                self.db_tx_status.configure(text="Unknown Board", text_color=("#F59E0B", "#F59E0B"))
        else:
            self.connection_badge.configure(text="Disconnected", fg_color="#7F1D1D", text_color="#FEE2E2")
            self.sidebar_status.configure(text="Disconnected", fg_color="#7F1D1D", text_color="#FEE2E2")
            self.db_tx_status.configure(text="Not Connected", text_color=("#EF4444", "#EF4444"))
            self.lora_active = False
            self.db_lora_status.configure(text="No Link", text_color=("#F59E0B", "#F59E0B"))

        self.port_stat.configure(text=port_text)
        self._update_port_selection_status()
        self._update_send_config_button()
        self._update_active_unit_display()

    def _send_quick(self, command: str) -> None:
        self._send_payload(command)

    def _send_time_sync(self) -> None:
        # Send local PC time to receiver over LoRa (no Wi-Fi dependency).
        now_local = datetime.datetime.now().replace(microsecond=0)
        self._send_payload(f"TIME:{now_local.strftime('%Y-%m-%d %H:%M:%S')}")

    def _send_custom(self) -> None:
        payload = self.command_var.get()
        if not payload:
            return
        self._send_payload(payload)

    def _send_payload(self, payload: str) -> None:
        if not self._is_transmitter_connected():
            messagebox.showwarning("Not Connected", "Connect to a transmitter first.")
            return

        try:
            wire_payload = payload if payload.endswith("\n") else f"{payload}\n"
            self.serial_service.send_text(wire_payload)
        except Exception as exc:
            messagebox.showerror("Send Error", str(exc))

    def _classify_line(self, line: str) -> str:
        if "TX>" in line:
            return "tx"
        if line.startswith("DATA:") or line.startswith("DATC:") or line.startswith("EVENT_FILE:"):
            return "data"
        if line.startswith("RSP:") or line.startswith("END:") or line.startswith("["):
            return "status"
        return "rx"

    def _refresh_log_widgets(self) -> None:
        """Refresh log-related widgets after one or more new lines were appended."""
        self.log_text.see("end")

        # Re-apply search highlight if active (debounced).
        if self.search_var.get().strip():
            self._schedule_search_highlight()

        self.tx_stat.configure(text=str(self.tx_count))
        self.rx_stat.configure(text=str(self.rx_count))
        self.db_tx_count.configure(text=str(self.tx_count))
        self.db_rx_count.configure(text=str(self.rx_count))
        self.db_log_count.configure(text=str(len(self.log_lines)))
        self.db_event_count.configure(text=str(self.session_events))
        self.db_last_message.configure(text=self.last_message)

    def _append_log(self, line: str, defer_ui: bool = False) -> None:
        self.log_lines.append(line)
        self.last_message = line
        tag = self._classify_line(line)

        inner: tk.Text = self.log_text._textbox  # type: ignore[attr-defined]
        inner.insert("end", line + "\n", tag)

        if tag == "tx":
            self.tx_count += 1
        else:
            self.rx_count += 1
            detected_role = self._role_from_banner(line)
            if detected_role is not None and self.connected_port != "-":
                self.detected_port_roles[self.connected_port] = detected_role
                if self.connected_role != detected_role:
                    self._set_connected_state(True, self.connected_port, role=detected_role)
            # Detect LoRa link from any receiver response
            if not self.lora_active and (
                line.startswith("RSP:") or line.startswith("DATA:")
                or line.startswith("DATC:") or line.startswith("END:")
            ):
                self.lora_active = True
                self.db_lora_status.configure(text="Link Active", text_color=("#22C55E", "#22C55E"))
            # Collect Connection Scan results
            if line.startswith("[SCAN_RESULT]:"):
                truck_id = line[14:].strip()
                if truck_id and truck_id not in self._scan_results:
                    self._scan_results.append(truck_id)
            # Track offload progress
            self._process_offload_message(line)

        # Each END:D marks a completed data offload from the receiver
        if line.startswith("END:D"):
            self.session_events += 1

        if not defer_ui:
            self._refresh_log_widgets()

    def _pump_messages(self) -> None:
        # While the top-level window is moving/resizing, defer heavy text/widget work.
        if time.time() < self._window_interacting_until:
            self.root.after(120, self._pump_messages)
            return

        processed = 0
        max_per_tick = 200
        while processed < max_per_tick and not self.serial_service.messages.empty():
            line = self.serial_service.messages.get_nowait()
            self._append_log(line, defer_ui=True)
            processed += 1
        if processed:
            self._refresh_log_widgets()
            self._update_active_unit_display()
        # Periodically refresh offload status display if active
        if self.offload_in_progress:
            now = time.time()
            if now >= self._next_offload_ui_update:
                self._update_offload_status_display()
                self._next_offload_ui_update = now + 0.25
        if not self.serial_service.messages.empty():
            self.root.after(10, self._pump_messages)
        else:
            self.root.after(100, self._pump_messages)

    def _clear_log(self) -> None:
        self.log_lines.clear()
        self.log_text.delete("1.0", "end")
        self.tx_count = 0
        self.rx_count = 0
        self.session_events = 0
        self.tx_stat.configure(text="0")
        self.rx_stat.configure(text="0")
        self.db_tx_count.configure(text="0")
        self.db_rx_count.configure(text="0")
        self.db_log_count.configure(text="0")
        self.db_event_count.configure(text="0")
        self.db_last_message.configure(text="No messages yet")
        # Reset offload stats when clearing log
        self.offload_in_progress = False
        self.offload_status = "Idle"
        self.offload_last_status = "Idle"
        self.offload_events_count = 0
        self._update_offload_status_display()
        self.search_var.set("")
        self._activate_search_placeholder()

    def _export_log(self) -> None:
        if not self.log_lines:
            messagebox.showinfo("No Data", "There is no log data to export.")
            return

        selected = filedialog.askdirectory(title="Choose Export Folder")
        if not selected:
            return

        output = export_text_log(self.log_lines, Path(selected))
        messagebox.showinfo("Export Complete", f"Saved log to:\n{output}")

    def _reset_offload_stats(self) -> None:
        """Reset offload tracking statistics for a new offload attempt."""
        self.offload_in_progress = True
        self.offload_start_time = time.time()
        self.offload_wifi_start_time = None
        self.offload_connected_time = None
        self.offload_data_start_time = None
        self.offload_end_time = None
        self.offload_connection_duration = None
        self.offload_transfer_duration = None
        self.offload_events_count = 0
        self.offload_status = "Connecting..."
        self.offload_last_status = "Offload started"
        self._offload_pending_event_name = None
        self._offload_pending_rows = []
        self._offload_lora_row_buf = ""
        self._offload_saved_files = []
        self._offload_session_dir = None
        self._offload_summary_file = None
        self._update_offload_status_display()

    def _get_offload_status_color(self) -> tuple[str, str]:
        return OFFLOAD_STATUS_COLORS.get(self.offload_status, ("#64748B", "#94A3B8"))

    def _update_offload_status_display(self) -> None:
        """Update the offload status card with current stats."""
        elapsed_text = "Elapsed: —"
        connection_text = "Connection: —"
        transfer_text = "Transfer: —"

        if not self.offload_in_progress:
            # Show last result when idle — preserve event count from last transfer
            self.offload_status_label.configure(
                text=self.offload_status,
                text_color=self._get_offload_status_color(),
            )

            if self.offload_end_time is not None:
                elapsed_text = f"Elapsed: {self.offload_end_time:.1f}s"
            if self.offload_connection_duration is not None:
                connection_text = f"Connection: {self.offload_connection_duration:.1f}s"
            if self.offload_transfer_duration is not None:
                transfer_text = f"Transfer: {self.offload_transfer_duration:.1f}s"

            self.offload_elapsed_label.configure(text=elapsed_text)
            self.offload_duration_label.configure(text=transfer_text)
            self.offload_connection_label.configure(text=connection_text)
            events_text = f"Events: {self.offload_events_count}" if self.offload_events_count else "Events: —"
            self.offload_events_label.configure(text=events_text)
            self.offload_message_label.configure(text=f"Last: {self.offload_last_status}")
            return

        # Update ongoing offload stats
        self.offload_status_label.configure(
            text=self.offload_status,
            text_color=self._get_offload_status_color(),
        )

        if self.offload_start_time:
            elapsed = time.time() - self.offload_start_time
            elapsed_text = f"Elapsed: {elapsed:.1f}s"

        # Connection duration (from wifi start to connected)
        if self.offload_wifi_start_time and self.offload_connected_time:
            conn_dur = self.offload_connected_time - self.offload_wifi_start_time
            connection_text = f"Connection: {conn_dur:.1f}s"

        # Transfer duration (from data start to now)
        if self.offload_data_start_time:
            transfer_dur = time.time() - self.offload_data_start_time
            transfer_text = f"Transfer: {transfer_dur:.1f}s"

        self.offload_elapsed_label.configure(text=elapsed_text)
        self.offload_connection_label.configure(text=connection_text)
        self.offload_duration_label.configure(text=transfer_text)

        # Events count
        self.offload_events_label.configure(text=f"Events: {self.offload_events_count}")

        # Last message
        self.offload_message_label.configure(text=f"Last: {self.offload_last_status}")

    def _process_offload_message(self, line: str) -> None:
        """Parse offload-related messages and update tracking stats."""
        # Offload start markers
        if "RSP:BEGIN_D" in line or "BEGIN_D" in line:
            self._reset_offload_stats()
        # Wi-Fi start
        elif "RSP:WIFI_START" in line or "WIFI_START" in line:
            if not self.offload_in_progress:
                self._reset_offload_stats()
            self.offload_wifi_start_time = time.time()
            self.offload_status = "Connecting..."
            self.offload_last_status = "Wi-Fi starting"
        # Wi-Fi connection attempts
        elif "RSP:WIFI_TRY:" in line or "[WIFI_TRY]" in line:
            self.offload_status = "Connecting..."
            try:
                if "RSP:WIFI_TRY:" in line:
                    network = line.split("RSP:WIFI_TRY:")[-1].strip().rstrip("]")
                else:
                    network = line.split("[WIFI_TRY]")[-1].strip()
                self.offload_last_status = f"Trying: {network}"
            except:
                self.offload_last_status = "Trying Wi-Fi..."
        # Wi-Fi connected successfully
        elif "RSP:WIFI_CONNECTED:" in line or "[RSP:WIFI_CONNECTED:" in line:
            self.offload_status = "Connecting..."
            try:
                ssid = line.split("WIFI_CONNECTED:")[-1].strip().rstrip("]")
                self.offload_last_status = f"WiFi connected: {ssid}"
            except:
                self.offload_last_status = "WiFi connected"
        # TCP server ready
        elif "RSP:WIFI_SERVER:" in line or "[WIFI_SERVER]" in line:
            self.offload_status = "Connecting..."
            self.offload_last_status = "TCP server ready"
        # Transmitter connected to TCP (data transfer beginning)
        elif "WIFI_TX_CONNECTED" in line or "RSP:WIFI_TX_CONNECTED" in line:
            self.offload_connected_time = time.time()
            self.offload_data_start_time = time.time()
            if self.offload_wifi_start_time:
                self.offload_connection_duration = self.offload_connected_time - self.offload_wifi_start_time
            self.offload_status = "Transferring..."
            self.offload_last_status = "Transmitter connected"
        # Wi-Fi countdown during TCP wait
        elif "RSP:WIFI_WAIT:" in line or "[RSP:WIFI_WAIT:" in line or "[WIFI_WAIT:" in line:
            try:
                countdown = line.split("WIFI_WAIT:")[-1].strip().rstrip("]").split()[0]
                self.offload_status = "Connecting..."
                self.offload_last_status = f"Waiting... {countdown}s"
            except:
                pass
        # Event file boundary marker — accept both direct receiver and transmitter-forwarded forms
        elif line.startswith("DATA:EVENT_FILE:") or line.startswith("EVENT_FILE:"):
            fname = line[16:].strip() if line.startswith("DATA:EVENT_FILE:") else line[11:].strip()
            self._save_pending_event()
            self._offload_pending_event_name = fname
            self._offload_pending_rows = []
            self._offload_lora_row_buf = ""
            if self.offload_status not in ("Transferring...",):
                if self.offload_data_start_time is None:
                    self.offload_data_start_time = time.time()
                self.offload_status = "Transferring..."
        # Partial LoRa chunk — accumulate until a final DATA: completes the row
        elif line.startswith("DATC:"):
            if self.offload_status != "Transferring...":
                self.offload_data_start_time = time.time()
                self.offload_status = "Transferring..."
            self._offload_lora_row_buf += line[5:]
        # Complete data row (Wi-Fi TCP full line, or final LoRa chunk)
        elif line.startswith("DATA:"):
            if self.offload_status != "Transferring...":
                self.offload_data_start_time = time.time()
                self.offload_status = "Transferring..."
            complete_row = self._offload_lora_row_buf + line[5:]
            self._offload_lora_row_buf = ""
            if self._offload_pending_event_name is not None:
                self._offload_pending_rows.append(complete_row)
            self.offload_events_count += 1
        # Raw data lines (numbers/CSV without DATA: prefix) during active transfer
        # Handles both digit-prefixed values and quoted timestamps ("Time not set",...)
        elif self.offload_in_progress and self.offload_data_start_time and line and (
            line[0].isdigit() or line.startswith('"')
        ):
            if self.offload_status != "Transferring...":
                self.offload_data_start_time = time.time()
                self.offload_status = "Transferring..."
            if self._offload_pending_event_name is not None:
                self._offload_pending_rows.append(line)
            self.offload_events_count += 1
        # TRANSFER summary message - arrives after END:D, update event count in last status
        elif "[TRANSFER]" in line:
            try:
                if "duration=" in line:
                    duration_str = line.split("duration=")[1].split()[0].strip().rstrip("ms")
                    self.offload_transfer_duration = float(duration_str) / 1000.0
                if "lines=" in line:
                    lines_str = line.split("lines=")[1].split()[0].strip()
                    self.offload_events_count = int(lines_str)
                    # Update last_status to replace the "0 events" with the real count
                    if self.offload_status == "Complete":
                        self.offload_last_status = self.offload_last_status.rsplit(",", 1)[0] + f", {self.offload_events_count} events)"
            except:
                pass
            self._update_offload_status_display()
        # Offload complete
        elif line.startswith("END:D"):
            if self.offload_in_progress:
                self._save_pending_event()  # flush last event file
                total_dur = time.time() - self.offload_start_time if self.offload_start_time else 0
                # Store timing info; real event count will be updated by [TRANSFER] shortly after
                self.offload_end_time = total_dur
                if self.offload_data_start_time:
                    self.offload_transfer_duration = time.time() - self.offload_data_start_time
                if self.offload_wifi_start_time and self.offload_connected_time:
                    self.offload_connection_duration = self.offload_connected_time - self.offload_wifi_start_time
                self.offload_in_progress = False
                self.offload_status = "Complete"
                saved_count = len(self._offload_saved_files)
                save_note = f" - {saved_count} file(s) saved" if saved_count else ""
                workbook_note = self._ensure_session_excel_summary() if saved_count else ""
                self.offload_last_status = (
                    f"Success ({total_dur:.1f}s, {self.offload_events_count} events){save_note}{workbook_note}"
                )
                self._update_offload_status_display()
        # Wi-Fi timeout
        elif "RSP:WIFI_TX_TIMEOUT" in line or "WIFI_TX_TIMEOUT" in line:
            if self.offload_in_progress:
                self._save_pending_event()  # flush last event file before ending
                total_dur = time.time() - self.offload_start_time if self.offload_start_time else 0
                self.offload_end_time = total_dur
                if self.offload_data_start_time:
                    self.offload_transfer_duration = time.time() - self.offload_data_start_time
                if self.offload_wifi_start_time and self.offload_connected_time:
                    self.offload_connection_duration = self.offload_connected_time - self.offload_wifi_start_time
                self.offload_in_progress = False
                saved_count = len(self._offload_saved_files)
                if saved_count > 0:
                    self.offload_status = "Complete"
                    workbook_note = self._ensure_session_excel_summary()
                    self.offload_last_status = (
                        f"Success without END:D ({total_dur:.1f}s, {self.offload_events_count} events)"
                        f" - {saved_count} file(s) saved{workbook_note}"
                    )
                else:
                    self.offload_status = "Timeout"
                    self.offload_last_status = f"Wi-Fi timeout ({total_dur:.1f}s)"
                self._update_offload_status_display()
        # Wi-Fi connection failed
        elif "RSP:WIFI_FAIL:" in line or "[WIFI_FAIL]" in line:
            if self.offload_in_progress:
                self._save_pending_event()  # flush last event file before ending
                if self.offload_start_time:
                    self.offload_end_time = time.time() - self.offload_start_time
                if self.offload_wifi_start_time and self.offload_connected_time:
                    self.offload_connection_duration = self.offload_connected_time - self.offload_wifi_start_time
                if self.offload_data_start_time:
                    self.offload_transfer_duration = time.time() - self.offload_data_start_time
                self.offload_in_progress = False
                self.offload_status = "Failed"
                saved_count = len(self._offload_saved_files)
                try:
                    if "[WIFI_FAIL]" in line:
                        network = line.split("[WIFI_FAIL]")[-1].strip()
                    else:
                        network = line.split("RSP:WIFI_FAIL:")[-1].strip().rstrip("]")
                    save_note = f" - {saved_count} file(s) saved" if saved_count else ""
                    workbook_note = self._ensure_session_excel_summary() if saved_count else ""
                    self.offload_last_status = f"Wi-Fi failed: {network}{save_note}{workbook_note}"
                except:
                    save_note = f" - {saved_count} file(s) saved" if saved_count else ""
                    workbook_note = self._ensure_session_excel_summary() if saved_count else ""
                    self.offload_last_status = f"Wi-Fi failed{save_note}{workbook_note}"
                self._update_offload_status_display()
        # Wi-Fi transmission layer failed (TCP, no profiles, or exhausted attempts)
        elif "[WIFI_TX_FAIL]" in line or "RSP:WIFI_TX_FAIL" in line:
            if self.offload_in_progress:
                self._save_pending_event()  # flush any pending event before ending
                total_dur = time.time() - self.offload_start_time if self.offload_start_time else 0
                self.offload_end_time = total_dur
                if self.offload_data_start_time:
                    self.offload_transfer_duration = time.time() - self.offload_data_start_time
                if self.offload_wifi_start_time and self.offload_connected_time:
                    self.offload_connection_duration = self.offload_connected_time - self.offload_wifi_start_time
                self.offload_in_progress = False
                self.offload_status = "Connection Failed"
                saved_count = len(self._offload_saved_files)
                try:
                    if "[WIFI_TX_FAIL]" in line:
                        reason = line.split("[WIFI_TX_FAIL]")[-1].strip()
                    else:
                        reason = line.split("RSP:WIFI_TX_FAIL")[-1].strip()
                    save_note = f" - {saved_count} file(s) saved" if saved_count else ""
                    workbook_note = self._ensure_session_excel_summary() if saved_count else ""
                    self.offload_last_status = (
                        f"{reason}{save_note}{workbook_note}" if reason else f"Wi-Fi unavailable{save_note}{workbook_note}"
                    )
                except:
                    save_note = f" - {saved_count} file(s) saved" if saved_count else ""
                    workbook_note = self._ensure_session_excel_summary() if saved_count else ""
                    self.offload_last_status = f"Wi-Fi unavailable{save_note}{workbook_note}"
                self._update_offload_status_display()
        # Fallback to LoRa
        elif "RSP:WIFI_FALLBACK_LORA" in line or "WIFI_FALLBACK_LORA" in line:
            self.offload_status = "Fallback LoRa"
            self.offload_last_status = "Switched to LoRa"
        # Receiver confirms stored events were cleared after offload
        elif "RSP:CLEAR_OK" in line or "CLEAR_OK" in line:
            if "Receiver cleared" not in self.offload_last_status:
                self.offload_last_status = f"{self.offload_last_status} | Receiver cleared"
            self._update_offload_status_display()
        # No data to transfer
        elif "RSP:NO_DATA" in line or "NO_DATA" in line:
            if self.offload_in_progress:
                self._save_pending_event()  # flush any pending event (though unlikely if no data)
                total_dur = time.time() - self.offload_start_time if self.offload_start_time else 0
                self.offload_end_time = total_dur
                if self.offload_wifi_start_time and self.offload_connected_time:
                    self.offload_connection_duration = self.offload_connected_time - self.offload_wifi_start_time
                self.offload_in_progress = False
                self.offload_status = "No Data"
                saved_count = len(self._offload_saved_files)
                save_note = f" - {saved_count} file(s) saved" if saved_count else ""
                workbook_note = self._ensure_session_excel_summary() if saved_count else ""
                self.offload_last_status = f"No events to transfer ({total_dur:.1f}s){save_note}{workbook_note}"
                self._update_offload_status_display()

    def _get_selected_setup_mask(self) -> int:
        mask = 0
        if self.apply_sensor_interval_var.get():
            mask |= SETUP_MASK_SENSOR_INTERVAL
        if self.apply_threshold_var.get():
            mask |= SETUP_MASK_THRESHOLD
        if self.apply_sample_rate_var.get():
            mask |= SETUP_MASK_SAMPLE_RATE
        if self.apply_duration_var.get():
            mask |= SETUP_MASK_DURATION
        if self.apply_truck_id_var.get():
            mask |= SETUP_MASK_TRUCK_ID
        if self.apply_description_var.get():
            mask |= SETUP_MASK_DESCRIPTION
        if self.apply_wifi_var.get():
            mask |= SETUP_MASK_WIFI
        return mask

    def _get_selected_setup_labels(self) -> list[str]:
        labels: list[str] = []
        if self.apply_sensor_interval_var.get():
            labels.append("Sensor Read Interval")
        if self.apply_threshold_var.get():
            labels.append("Event Trigger Threshold")
        if self.apply_sample_rate_var.get():
            labels.append("Strain Gauge Poll Rate")
        if self.apply_duration_var.get():
            labels.append("Event Capture Duration")
        if self.apply_truck_id_var.get():
            labels.append("Truck ID")
        if self.apply_description_var.get():
            labels.append("Description")
        if self.apply_wifi_var.get():
            labels.append("Wi-Fi Network")
        return labels

    def _update_send_config_button(self) -> None:
        if self.send_config_button is None:
            return

        ready_to_send = self._get_selected_setup_mask() != 0 and self._is_transmitter_connected()
        if ready_to_send:
            self.send_config_button.configure(
                state="normal",
                fg_color=WABASH_BLUE,
                hover_color=WABASH_BLUE_HOVER,
            )
        else:
            self.send_config_button.configure(
                state="disabled",
                fg_color=BTN_GREY,
                hover_color=BTN_GREY,
            )

    def _on_setup_selection_changed(self, *_args: object) -> None:
        self._update_send_config_button()

    def _set_all_setup_selection(self, selected: bool) -> None:
        for var in self._setup_apply_vars:
            var.set(selected)

    def _role_from_banner(self, line: str) -> str | None:
        lowered = line.lower()
        if "heltec lora transmitter bridge" in lowered:
            return "transmitter"
        if "heltec capstone receiver" in lowered:
            return "receiver"
        return None

    def _is_transmitter_connected(self) -> bool:
        return self.serial_service.is_connected and self.connected_role == "transmitter"

    def _update_port_selection_status(self) -> None:
        selected_port = self.port_var.get().strip()
        role = self.detected_port_roles.get(selected_port, "unknown") if selected_port and selected_port != "No ports found" else "none"

        if selected_port and selected_port != "No ports found":
            if role == "transmitter":
                text = f"Selected Board: Transmitter on {selected_port}"
                conn_text = f"Board Detection: Transmitter on {selected_port}"
            elif role == "receiver":
                text = f"Selected Board: Receiver on {selected_port}"
                conn_text = f"Board Detection: Receiver on {selected_port}"
            else:
                text = f"Selected Board: Unknown on {selected_port}"
                conn_text = f"Board Detection: Unknown on {selected_port}"
        else:
            text = "Selected Board: None"
            conn_text = "Board Detection: Unknown"

        if hasattr(self, "sidebar_board_label"):
            self.sidebar_board_label.configure(text=text)
        if hasattr(self, "port_detect_label"):
            self.port_detect_label.configure(text=conn_text)

    def _probe_port_role(self, port: str, baudrate: int = 115200) -> str:
        try:
            with serial.Serial(port=port, baudrate=baudrate, timeout=0.2, write_timeout=0.2) as ser:
                try:
                    ser.reset_input_buffer()
                    ser.setDTR(False)
                    ser.setRTS(False)
                    time.sleep(0.05)
                    ser.setDTR(True)
                except Exception:
                    pass

                deadline = time.monotonic() + 2.0
                while time.monotonic() < deadline:
                    raw = ser.readline()
                    if not raw:
                        continue
                    line = raw.decode("utf-8", errors="replace").strip()
                    detected_role = self._role_from_banner(line)
                    if detected_role is not None:
                        return detected_role
        except Exception:
            return "unknown"

        return "unknown"

    def _auto_detect_board(self) -> None:
        ports = [port.device for port in list_ports.comports()]
        if not ports:
            messagebox.showwarning("No Ports", "No serial boards were detected.")
            return

        found_transmitter: str | None = None
        found_receiver: str | None = None
        for port in ports:
            role = self._probe_port_role(port)
            self.detected_port_roles[port] = role
            if role == "transmitter" and found_transmitter is None:
                found_transmitter = port
            elif role == "receiver" and found_receiver is None:
                found_receiver = port

        self._refresh_ports()

        if found_transmitter is not None:
            self.port_var.set(found_transmitter)
            self._append_log(f"[AUTO_DETECT] Transmitter detected on {found_transmitter}")
            if self._connect(suppress_warnings=True):
                self._append_log(f"[AUTO_DETECT] Auto-connected to {found_transmitter}")
            else:
                messagebox.showwarning("Auto Connect", f"Transmitter found on {found_transmitter}, but connection failed.")
            return

        if found_receiver is not None:
            self.port_var.set(found_receiver)
            self._append_log(f"[AUTO_DETECT] Receiver detected on {found_receiver}")
            messagebox.showwarning("Receiver Detected", "A receiver was detected, but the interface requires a transmitter connection.")
            return

        self._append_log("[AUTO_DETECT] Unable to determine board type for detected ports.")
        messagebox.showwarning("Auto Detect", "Board(s) were found, but the role could not be verified from serial output.")

    def _auto_connect_tick(self) -> None:
        if self._is_transmitter_connected():
            self.root.after(3000, self._auto_connect_tick)
            return

        if self._auto_connect_inflight:
            self.root.after(1200, self._auto_connect_tick)
            return

        self._auto_connect_inflight = True
        worker = threading.Thread(target=self._scan_and_connect_transmitter_worker, daemon=True)
        worker.start()

    def _scan_and_connect_transmitter_worker(self) -> None:
        found_port: str | None = None
        try:
            ports = [port.device for port in list_ports.comports()]
            for port in ports:
                role = self._probe_port_role(port)
                self.detected_port_roles[port] = role
                if role == "transmitter":
                    found_port = port
                    break
        finally:
            self.root.after(0, lambda p=found_port: self._complete_auto_connect_scan(p))

    def _complete_auto_connect_scan(self, found_port: str | None) -> None:
        try:
            self._refresh_ports()
            if found_port is not None and not self._is_transmitter_connected():
                self.port_var.set(found_port)
                if self._connect(suppress_warnings=True):
                    self._append_log(f"[AUTO_CONNECT] Connected to transmitter on {found_port}")
        finally:
            self._auto_connect_inflight = False
            self.root.after(3000, self._auto_connect_tick)

    def _build_unit_setup_page(self) -> None:
        page = ctk.CTkScrollableFrame(
            self.page_container,
            corner_radius=0,
            fg_color="transparent",
            scrollbar_button_color=BTN_GREY,
            scrollbar_button_hover_color=BTN_GREY_HOVER,
        )
        page.grid_columnconfigure(0, weight=1)
        self.pages["Unit Setup"] = page
        page.bind("<Configure>", lambda _e: self._schedule_unit_setup_wrap(), add="+")

        header = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        header.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        ctk.CTkLabel(header, text="Unit Configuration", font=ctk.CTkFont(size=22, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=18, pady=14
        )

        info_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        info_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        info_card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            info_card,
            text="Truck Identification",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(14, 6))

        ctk.CTkLabel(info_card, text="Truck ID", text_color=("#475569", "#94A3B8")).grid(
            row=1, column=0, sticky="w", padx=16, pady=6
        )
        ctk.CTkEntry(
            info_card,
            textvariable=self.truck_id_var,
            placeholder_text="Truck ID",
        ).grid(row=1, column=1, sticky="ew", padx=(6, 16), pady=6)

        ctk.CTkLabel(info_card, text="Description", text_color=("#475569", "#94A3B8")).grid(
            row=2, column=0, sticky="w", padx=16, pady=(6, 14)
        )
        ctk.CTkEntry(
            info_card,
            textvariable=self.description_var,
            placeholder_text="Description",
        ).grid(row=2, column=1, sticky="ew", padx=(6, 16), pady=(6, 14))

        config_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        config_card.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        config_card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            config_card,
            text="Sensor Configuration",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(14, 6))

        ctk.CTkLabel(config_card, text="Sensor Read Interval (ms)", text_color=("#475569", "#94A3B8")).grid(
            row=1, column=0, sticky="w", padx=16, pady=6
        )
        ctk.CTkOptionMenu(
            config_card,
            values=["50", "100", "200", "500"],
            variable=self.sensor_interval_var,
            fg_color=WABASH_BLUE,
            button_color=WABASH_BLUE_HOVER,
            button_hover_color=WABASH_BLUE_HOVER,
        ).grid(row=1, column=1, sticky="ew", padx=(6, 16), pady=6)

        ctk.CTkLabel(config_card, text="Strain Gauge Poll Rate (Hz)", text_color=("#475569", "#94A3B8")).grid(
            row=2, column=0, sticky="w", padx=16, pady=6
        )
        ctk.CTkOptionMenu(
            config_card,
            values=["10", "20"],
            variable=self.lab_sample_rate_var,
            fg_color=WABASH_BLUE,
            button_color=WABASH_BLUE_HOVER,
            button_hover_color=WABASH_BLUE_HOVER,
        ).grid(row=2, column=1, sticky="ew", padx=(6, 16), pady=6)

        ctk.CTkLabel(config_card, text="Event Trigger Threshold (g's)", text_color=("#475569", "#94A3B8")).grid(
            row=3, column=0, sticky="w", padx=16, pady=6
        )
        ctk.CTkEntry(
            config_card,
            textvariable=self.event_trigger_threshold_var,
            placeholder_text="Example: 2.0",
        ).grid(row=3, column=1, sticky="ew", padx=(6, 16), pady=6)

        ctk.CTkLabel(config_card, text="Event Capture Duration (ms)", text_color=("#475569", "#94A3B8")).grid(
            row=4, column=0, sticky="w", padx=16, pady=(6, 14)
        )
        ctk.CTkEntry(
            config_card,
            textvariable=self.event_duration_var,
            placeholder_text="Example: 2000",
        ).grid(row=4, column=1, sticky="ew", padx=(6, 16), pady=(6, 14))

        wifi_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        wifi_card.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        wifi_card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            wifi_card,
            text="Wi-Fi Offload Network",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(14, 6))

        ctk.CTkLabel(wifi_card, text="SSID", text_color=("#475569", "#94A3B8")).grid(
            row=1, column=0, sticky="w", padx=16, pady=6
        )
        ctk.CTkEntry(
            wifi_card,
            textvariable=self.wifi1_ssid_var,
            placeholder_text="Wi-Fi SSID",
        ).grid(row=1, column=1, sticky="ew", padx=(6, 16), pady=6)

        ctk.CTkLabel(wifi_card, text="Password", text_color=("#475569", "#94A3B8")).grid(
            row=2, column=0, sticky="w", padx=16, pady=(6, 14)
        )
        ctk.CTkEntry(
            wifi_card,
            textvariable=self.wifi1_password_var,
            placeholder_text="Wi-Fi Password",
            show="*",
        ).grid(row=2, column=1, sticky="ew", padx=(6, 16), pady=(6, 14))

        desc_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        desc_card.grid(row=5, column=0, sticky="ew", pady=(0, 12))
        desc_card.grid_columnconfigure(0, weight=1)
        desc_card.bind("<Configure>", lambda _e: self._schedule_unit_setup_wrap(), add="+")
        self.unit_setup_desc_card = desc_card
        self.unit_setup_help_label = ctk.CTkLabel(
            desc_card,
            text="Select the fields to include above. Unselected values are left unchanged on the receiver. Selecting Wi-Fi with blank SSID/password clears the stored Wi-Fi network. To send a configuration, at least one field must be selected and you must be connected to a transmitter.",
            wraplength=800,
            justify="left",
            text_color=("#334155", "#CBD5E1"),
            font=ctk.CTkFont(size=12),
        )
        self.unit_setup_help_label.grid(row=0, column=0, sticky="ew", padx=16, pady=12)

        update_card = ctk.CTkFrame(page, corner_radius=14, fg_color=(CARD_LIGHT, CARD_DARK))
        update_card.grid(row=4, column=0, sticky="ew", pady=(0, 12))
        update_card.grid_columnconfigure((0, 1, 2), weight=1)

        ctk.CTkLabel(
            update_card,
            text="Include In This Update",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=16, pady=(14, 10))

        update_options = [
            ("Truck ID", self.apply_truck_id_var),
            ("Description", self.apply_description_var),
            ("Sensor Interval", self.apply_sensor_interval_var),
            ("Trigger Threshold", self.apply_threshold_var),
            ("Poll Rate", self.apply_sample_rate_var),
            ("Capture Duration", self.apply_duration_var),
            ("Wi-Fi Network", self.apply_wifi_var),
        ]
        for idx, (label, var) in enumerate(update_options):
            row = 1 + (idx // 3)
            col = idx % 3
            ctk.CTkCheckBox(
                update_card,
                text=label,
                variable=var,
                onvalue=True,
                offvalue=False,
                fg_color=WABASH_BLUE,
                hover_color=WABASH_BLUE_HOVER,
            ).grid(row=row, column=col, sticky="w", padx=16, pady=(0, 10 if idx < 6 else 14))

        button_bar = ctk.CTkFrame(update_card, fg_color="transparent")
        button_bar.grid(row=4, column=0, columnspan=3, sticky="w", padx=16, pady=(0, 14))

        ctk.CTkButton(
            button_bar,
            text="Select All",
            width=110,
            command=lambda: self._set_all_setup_selection(True),
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
        ).grid(row=0, column=0, padx=(0, 8))

        ctk.CTkButton(
            button_bar,
            text="Deselect All",
            width=110,
            command=lambda: self._set_all_setup_selection(False),
            fg_color=BTN_GREY,
            hover_color=BTN_GREY_HOVER,
        ).grid(row=0, column=1)

        self.send_config_button = ctk.CTkButton(
            page,
            text="Send Configuration",
            command=self._send_unit_config,
            fg_color=WABASH_BLUE,
            hover_color=WABASH_BLUE_HOVER,
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40,
            state="disabled",
        )
        self.send_config_button.grid(row=6, column=0, sticky="ew")
        self._update_send_config_button()
        self.root.after(0, self._update_unit_setup_wrap)

    def _schedule_unit_setup_wrap(self) -> None:
        """Debounce wraplength recalculation so it only runs once after resize settles."""
        if self._resizing:
            return
        if hasattr(self, "_wrap_job") and self._wrap_job is not None:
            self.root.after_cancel(self._wrap_job)
        self._wrap_job = self.root.after(120, self._run_unit_setup_wrap)

    def _run_unit_setup_wrap(self) -> None:
        self._wrap_job = None
        self._update_unit_setup_wrap()

    def _update_unit_setup_wrap(self) -> None:
        if self.unit_setup_help_label is None:
            return

        try:
            if self.unit_setup_desc_card is not None:
                container_width = self.unit_setup_desc_card.winfo_width()
            else:
                page = self.pages.get("Unit Setup")
                if page is None:
                    return
                container_width = page.winfo_width()

            if container_width <= 1:
                return

            # Account for label side padding so wrapping matches the card's visible width.
            self.unit_setup_help_label.configure(wraplength=max(380, container_width - 36))
        except Exception:
            return

    def _load_app_settings(self) -> None:
        # Ensure app directory exists on startup
        try:
            _APP_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        
        try:
            if _SETTINGS_FILE.exists():
                with _SETTINGS_FILE.open("r", encoding="utf-8") as f:
                    settings = json.load(f)
                if settings.get("data_path"):
                    self.data_path_var.set(settings["data_path"])
        except Exception:
            pass

    def _save_app_settings(self) -> None:
        try:
            _APP_DIR.mkdir(parents=True, exist_ok=True)
            with _SETTINGS_FILE.open("w", encoding="utf-8") as f:
                json.dump({"data_path": self.data_path_var.get()}, f, indent=2)
        except Exception:
            pass

    def _browse_data_path(self) -> None:
        folder = filedialog.askdirectory(title="Choose Data Save Location")
        if folder:
            self.data_path_var.set(folder)
            self._save_app_settings()
            self._refresh_data_viewer()
    
    def _open_data_folder(self) -> None:
        """Open the current data folder in file explorer."""
        try:
            data_path = Path(self.data_path_var.get())
            data_path.mkdir(parents=True, exist_ok=True)
            if sys.platform == "win32":
                os.startfile(str(data_path))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(data_path)])
            else:
                subprocess.Popen(["xdg-open", str(data_path)])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder: {e}")

    def _save_pending_event(self) -> None:
        """Flush buffered event rows to a local CSV file and reset the buffer."""
        name = self._offload_pending_event_name
        rows = list(self._offload_pending_rows)
        # Clear state immediately to be safe against any re-entrant call
        self._offload_pending_event_name = None
        self._offload_pending_rows = []
        self._offload_lora_row_buf = ""

        if not name or not rows:
            return

        truck_id = self.truck_id_var.get().strip() or "Unknown"
        safe_truck = (
            "".join(c if c.isalnum() or c in " _-" else "_" for c in truck_id).strip()
            or "Unknown"
        )

        try:
            base = Path(self.data_path_var.get())
            truck_dir = base / safe_truck
            truck_dir.mkdir(parents=True, exist_ok=True)

            # Create one unique folder per offload session under the Truck ID directory.
            if self._offload_session_dir is None:
                ts_source = self.offload_start_time if self.offload_start_time is not None else time.time()
                ts_name = datetime.datetime.fromtimestamp(ts_source).strftime("%Y-%m-%d_%H-%M-%S")
                session_dir = truck_dir / ts_name
                if session_dir.exists():
                    suffix = 2
                    while (truck_dir / f"{ts_name}_{suffix}").exists():
                        suffix += 1
                    session_dir = truck_dir / f"{ts_name}_{suffix}"
                session_dir.mkdir(parents=True, exist_ok=True)
                self._offload_session_dir = session_dir

            safe_fname = name.replace("/", "_").replace("\\", "_")
            dest = self._offload_session_dir / safe_fname
            if dest.exists():
                stem = dest.stem
                suffix = dest.suffix
                idx = 2
                while True:
                    candidate = self._offload_session_dir / f"{stem}_{idx}{suffix}"
                    if not candidate.exists():
                        dest = candidate
                        break
                    idx += 1
            with dest.open("w", encoding="utf-8", newline="") as f:
                for row in rows:
                    f.write(row + "\n")
            self._offload_saved_files.append(dest)
        except Exception as exc:
            # Defer log call to avoid re-entering _process_offload_message
            self.root.after(0, lambda msg=f"[SAVE_ERR] {name}: {exc}": self._append_log(msg))

    def _parse_event_row(self, row: str) -> list[str]:
        """Parse one event CSV row while preserving quoted timestamps and text fields."""
        try:
            parsed = next(csv.reader([row]))
        except Exception:
            parsed = row.split(",")
        return [cell.strip() for cell in parsed]

    def _coerce_excel_value(self, value: str) -> object:
        """Convert numeric-looking CSV values to real Excel numbers."""
        text = value.strip()
        if not text:
            return ""
        try:
            return int(text)
        except ValueError:
            pass
        try:
            return float(text)
        except ValueError:
            return text

    def _build_session_excel_summary(self) -> Path | None:
        """Combine all saved event files in this offload session into one Excel workbook."""
        if not self._offload_saved_files or self._offload_session_dir is None:
            return None

        try:
            from openpyxl import Workbook
        except Exception:
            return None

        rows_for_excel: list[tuple[int, list[str]]] = []
        max_extra_cols = 0

        for event_index, event_path in enumerate(self._offload_saved_files, start=1):
            try:
                text = event_path.read_text(encoding="utf-8")
            except Exception:
                continue

            for raw in text.splitlines():
                line = raw.strip()
                if not line:
                    continue
                fields = self._parse_event_row(line)
                if len(fields) < 8:
                    fields.extend([""] * (8 - len(fields)))
                extra = max(0, len(fields) - 8)
                if extra > max_extra_cols:
                    max_extra_cols = extra
                rows_for_excel.append((event_index, fields))

        if not rows_for_excel:
            return None

        headers = [
            "Event #",
            "Time Stamp",
            "Temp",
            "RH",
            "Accelx",
            "Accely",
            "Accelz",
            "Strain",
        ]

        sample_group_count = (max_extra_cols + 3) // 4
        for sample_idx in range(2, sample_group_count + 2):
            headers.extend([
                f"Accelx_{sample_idx}",
                f"Accely_{sample_idx}",
                f"Accelz_{sample_idx}",
                f"Strain_{sample_idx}",
            ])

        wb = Workbook()
        ws = wb.active
        ws.title = "Offload Data"

        truck_id = self.truck_id_var.get().strip() or "Unknown"
        truck_description = self.description_var.get().strip() or ""
        ws.append(["Truck ID:", truck_id])
        ws.append(["Truck Description:", truck_description])
        ws.append([])
        ws.append(headers)

        for event_index, fields in rows_for_excel:
            row_values: list[object] = [event_index]
            padded_fields = fields + [""] * (8 + max_extra_cols - len(fields))
            row_values.append(padded_fields[0])
            row_values.extend(self._coerce_excel_value(value) for value in padded_fields[1 : 8 + max_extra_cols])
            ws.append(row_values)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 28)

        out_path = self._offload_session_dir / "combined_offload.xlsx"
        if out_path.exists():
            idx = 2
            while (self._offload_session_dir / f"combined_offload_{idx}.xlsx").exists():
                idx += 1
            out_path = self._offload_session_dir / f"combined_offload_{idx}.xlsx"

        wb.save(out_path)
        return out_path

    def _ensure_session_excel_summary(self) -> str:
        """Create the combined workbook once per offload session and return a status suffix."""
        if self._offload_summary_file is not None:
            return f" - workbook: {self._offload_summary_file.name}"

        try:
            out_path = self._build_session_excel_summary()
        except Exception as exc:
            self.root.after(0, lambda msg=f"[XLSX_ERR] {exc}": self._append_log(msg))
            return ""

        if out_path is None:
            return ""

        self._offload_summary_file = out_path
        self.root.after(0, self._refresh_data_viewer)
        return f" - workbook: {out_path.name}"

    def _send_unit_config(self) -> None:
        if not self._is_transmitter_connected():
            messagebox.showwarning("Not Connected", "Connect to a transmitter first.")
            return

        try:
            setup_mask = self._get_selected_setup_mask()
            if setup_mask == 0:
                messagebox.showwarning("No Fields Selected", "Select at least one field to include in this update.")
                return

            interval = 0
            if self.apply_sensor_interval_var.get():
                interval = int(self.sensor_interval_var.get().strip())

            threshold = 0.0
            if self.apply_threshold_var.get():
                threshold = float(self.event_trigger_threshold_var.get().strip())

            sample_rate = 0
            if self.apply_sample_rate_var.get():
                sample_rate = int(self.lab_sample_rate_var.get().strip())

            duration = 0
            if self.apply_duration_var.get():
                duration = int(self.event_duration_var.get().strip())

            truck_id = self.truck_id_var.get().replace(";", " ").replace("=", " ").replace("\n", " ").replace("\r", " ").strip()
            description = self.description_var.get().replace(";", " ").replace("=", " ").replace("\n", " ").replace("\r", " ").strip()

            wifi_ssid = self.wifi1_ssid_var.get().replace(";", " ").replace("=", " ").replace("\n", " ").replace("\r", " ").strip()
            wifi_password = self.wifi1_password_var.get().replace(";", " ").replace("=", " ").replace("\n", " ").replace("\r", " ").strip()

            if self.apply_wifi_var.get() and wifi_password and not wifi_ssid:
                messagebox.showwarning(
                    "Wi-Fi Setup Error",
                    "Wi-Fi password is set, but the SSID is empty.",
                )
                return

            wifi_fields = [
                f"w0s={wifi_ssid}",
                f"w0p={wifi_password}",
                "w1s=",
                "w1p=",
                "w2s=",
                "w2p=",
            ]

            packet = (
                f"SETUP:m={setup_mask};si={interval};thr={threshold};sr={sample_rate};dur={duration};"
                f"tid={truck_id};desc={description};"
                + ";".join(wifi_fields)
            )

            # Transmitter forwards SETUP in one LoRa frame, so keep payload conservative.
            if len(packet.encode("utf-8")) > 220:
                messagebox.showwarning(
                    "Setup Too Large",
                    "Configuration is too large for a single LoRa setup packet. "
                    "Shorten SSID/password/description values and try again.",
                )
                return

            wire_payload = f"{packet}\n"
            self.serial_service.send_text(wire_payload)

            wifi_configured = 1 if (self.apply_wifi_var.get() and wifi_ssid) else 0
            updated_fields = ", ".join(self._get_selected_setup_labels())

            messagebox.showinfo(
                "Configuration Sent",
                f"Unit configuration sent:\n\n"
                f"Fields Updated: {updated_fields}\n"
                f"Wi-Fi Network Saved: {wifi_configured}"
            )
        except ValueError:
            messagebox.showerror("Invalid Input", "Selected numeric fields must contain valid numeric values.")
        except Exception as exc:
            messagebox.showerror("Send Error", str(exc))

    def _update_active_unit_display(self) -> None:
        """Refresh the Active Unit panel on the dashboard from Unit Setup values."""
        if self.truck_id_var.get().strip():
            tid = self.truck_id_var.get().strip()
            self.db_unit_id.configure(text=tid, text_color=("#1E293B", "#E2E8F0"))
            if self.description_var.get().strip():
                self.db_unit_desc.configure(text=self.description_var.get().strip())
            else:
                self.db_unit_desc.configure(text="No description set")
        else:
            self.db_unit_id.configure(text="No unit configured", text_color=("#64748B", "#9CA3AF"))
            self.db_unit_desc.configure(text="")

    def _load_known_units(self) -> None:
        self.known_truck_ids.clear()
        # File format: "truck_id|YYYY-MM-DD HH:MM" per line (legacy: bare truck_id)
        last_seen_map: dict[str, str] = {}
        try:
            _APP_DIR.mkdir(parents=True, exist_ok=True)
            if _KNOWN_TRUCKS_FILE.exists():
                for raw_line in _KNOWN_TRUCKS_FILE.read_text(encoding="utf-8").splitlines():
                    parts = raw_line.strip().split("|", 1)
                    truck_id = parts[0].strip()
                    if truck_id:
                        self.known_truck_ids.add(truck_id)
                        last_seen_map[truck_id] = parts[1].strip() if len(parts) == 2 else "-"
        except Exception:
            self.known_truck_ids.clear()

        for truck_id in sorted(self.known_truck_ids):
            if truck_id not in self.units:
                self.units[truck_id] = {
                    "truck_id": truck_id,
                    "last_seen": last_seen_map.get(truck_id, "-"),
                    "events": 0,
                    "status": "Known",
                }

    def _save_known_units(self) -> None:
        try:
            _APP_DIR.mkdir(parents=True, exist_ok=True)
            lines = []
            for truck_id in sorted(self.known_truck_ids):
                last_seen = self.units.get(truck_id, {}).get("last_seen", "-")
                lines.append(f"{truck_id}|{last_seen}")
            _KNOWN_TRUCKS_FILE.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")
        except Exception:
            pass

    def _delete_known_unit(self, truck_id: str) -> None:
        if not messagebox.askyesno(
            "Delete Known Unit",
            f"Delete truck ID '{truck_id}' from known units?",
        ):
            return

        self.known_truck_ids.discard(truck_id)
        self._save_known_units()
        if truck_id in self.units:
            del self.units[truck_id]
        self._refresh_unit_list()

    def _register_unit(self, truck_id: str) -> None:
        """Add or update a unit in the fleet registry and refresh the table."""
        import datetime
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        self.known_truck_ids.add(truck_id)

        if truck_id in self.units:
            self.units[truck_id]["last_seen"] = now
            self.units[truck_id]["events"] += 1
            self.units[truck_id]["status"] = "Active"
        else:
            self.units[truck_id] = {
                "truck_id": truck_id,
                "last_seen": now,
                "events": 1,
                "status": "Active",
            }
        self._save_known_units()
        self._refresh_unit_list()

    def _refresh_unit_list(self) -> None:
        """Rebuild the fleet table rows, applying the current filter."""
        for row_widgets in self._fleet_rows:
            for w in row_widgets:
                w.destroy()
        self._fleet_rows.clear()

        for truck_id in self.known_truck_ids:
            if truck_id not in self.units:
                self.units[truck_id] = {
                    "truck_id": truck_id,
                    "last_seen": "-",
                    "events": 0,
                    "status": "Known",
                }

        term = self.unit_filter_var.get().strip().lower()
        units = [u for u in self.units.values() if not term or term in u["truck_id"].lower()]
        units.sort(key=lambda u: (u["status"] != "Active", u["truck_id"]))

        if not units:
            self.fleet_empty_label.grid(row=1, column=0, columnspan=5, pady=30)
            return

        self.fleet_empty_label.grid_remove()
        fleet_col_mins = (120, 98, 64, 76, 74)
        fleet_col_weights = (4, 3, 2, 2, 2)
        row_colors = (("#EAF1FB", "#152946"), ("#DEE9F8", "#1A3153"))

        for i, unit in enumerate(units):
            row_widgets: list[ctk.CTkBaseClass] = []
            # Row frame placed in single scroll column → fills full width → smooth background.
            row_frame = ctk.CTkFrame(
                self.fleet_scroll,
                corner_radius=6,
                fg_color=row_colors[i % 2],
            )
            row_frame.grid(row=i + 2, column=0, sticky="ew", pady=2)
            # Same column layout as fleet_thead → columns align perfectly.
            for col, (min_w, weight) in enumerate(zip(fleet_col_mins, fleet_col_weights)):
                row_frame.grid_columnconfigure(col, minsize=min_w, weight=weight)
            row_widgets.append(row_frame)

            for col, text in enumerate([unit["truck_id"], unit["last_seen"], str(unit["events"]), unit["status"]]):
                lbl = ctk.CTkLabel(
                    row_frame, text=text,
                    font=ctk.CTkFont(size=12),
                    text_color=("#1E293B", "#E2E8F0"),
                    anchor="w",
                )
                lbl.grid(row=0, column=col, sticky="ew", padx=12, pady=8)
                row_widgets.append(lbl)

            del_btn = ctk.CTkButton(
                row_frame,
                text="Delete",
                width=68,
                height=26,
                fg_color="#B91C1C",
                hover_color="#991B1B",
                text_color="#FEE2E2",
                font=ctk.CTkFont(size=11, weight="bold"),
                command=lambda tid=unit["truck_id"]: self._delete_known_unit(tid),
            )
            del_btn.grid(row=0, column=4, sticky="e", padx=(0, 8), pady=6)
            row_widgets.append(del_btn)
            self._fleet_rows.append(row_widgets)

    def _run_connection_scan(self) -> None:
        if not self._is_transmitter_connected():
            messagebox.showwarning("Not Connected", "Connect to a transmitter first.")
            return
        self._scan_results.clear()
        self.scan_button.configure(text="Scanning...", state="disabled")
        self._clear_disc_table()
        self.disc_empty_label.configure(text="Scanning for units...")
        self._send_payload("SCAN")
        self.root.after(3000, self._finish_connection_scan)

    def _finish_connection_scan(self) -> None:
        self.scan_button.configure(text="Unit Discover", state="normal")
        if self._scan_results:
            self._refresh_disc_table()
        else:
            self.disc_empty_label.configure(
                text="No units found.\nPress Unit Discover to scan.")
            self.disc_empty_label.grid(row=0, column=0, columnspan=2, pady=20)

    def _clear_disc_table(self) -> None:
        for row_widgets in self._disc_rows:
            for w in row_widgets:
                w.destroy()
        self._disc_rows.clear()
        self.disc_empty_label.grid(row=0, column=0, columnspan=2, pady=20)

    def _refresh_disc_table(self) -> None:
        self._clear_disc_table()
        self.disc_empty_label.grid_forget()
        self.disc_scroll.grid_columnconfigure(0, weight=1)
        for idx, unit_id in enumerate(self._scan_results):
            bg = (CARD_LIGHT, CARD_DARK) if idx % 2 == 0 else ('#F1F5F9', '#1E3050')
            row_frame = ctk.CTkFrame(self.disc_scroll, corner_radius=6,
                                     fg_color=bg)
            row_frame.grid(row=idx, column=0, sticky='ew', pady=2)
            row_frame.grid_columnconfigure(0, weight=1)
            lbl = ctk.CTkLabel(row_frame, text=unit_id,
                               font=ctk.CTkFont(size=13))
            lbl.grid(row=0, column=0, sticky='w', padx=10, pady=6)
            sel_btn = ctk.CTkButton(
                row_frame, text='Select', width=70,
                fg_color=WABASH_BLUE, hover_color=WABASH_BLUE_HOVER,
                command=lambda uid=unit_id: self._select_scanned_unit(uid))
            sel_btn.grid(row=0, column=1, padx=(4, 10), pady=6)
            self._disc_rows.append([row_frame])

    def _select_scanned_unit(self, unit_id: str) -> None:
        self.truck_id_var.set(unit_id)
        self.apply_truck_id_var.set(True)
        self._update_active_unit_display()
        self._register_unit(unit_id)
        # Enable unit action buttons
        for btn in (self.btn_request, self.btn_tare, self.btn_timesync):
            btn.configure(state="normal",
                          fg_color=WABASH_BLUE,
                          hover_color=WABASH_BLUE_HOVER)
        self._append_log(f"[Status] Active unit set to: {unit_id}")

    def _on_close(self) -> None:
        self.serial_service.disconnect()
        self.root.destroy()

    def _on_root_configure(self, event: object) -> None:
        """Debounce root window resize: freeze updates during drag, flush once settled."""
        if getattr(event, "widget", None) is not self.root:
            return
        self._window_interacting_until = time.time() + 0.12
        current_size = (event.width, event.height)
        if current_size == self._last_size:
            return  # only position changed (window drag), skip entirely
        self._last_size = current_size
        self._resizing = True
        if self._resize_job is not None:
            self.root.after_cancel(self._resize_job)
        self._resize_job = self.root.after(80, self._on_resize_settled)

    def _on_resize_settled(self) -> None:
        self._resize_job = None
        self._resizing = False
        self._update_unit_setup_wrap()

